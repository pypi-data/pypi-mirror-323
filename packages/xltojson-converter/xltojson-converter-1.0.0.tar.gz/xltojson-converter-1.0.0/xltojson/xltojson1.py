from openpyxl import load_workbook
import json

class ExcelToJson:
    def __init__(self, excel_file, json_file):
        self.excel_file = excel_file
        self.json_file = json_file
        self.data = {}

    def read_excel(self):
        """Reads the Excel file and loads the data into a hierarchical structure."""
        wb = load_workbook(self.excel_file)
        sheet = wb.active

        # Initialize variables for hierarchy tracking
        current_parent = None
        current_child = None
        current_subchild = None
        collecting_options = False

        for row in sheet.iter_rows(min_row=1, values_only=True):
            # Check for Parent (Column 1)
            if row[0]:
                current_parent = row[0]
                if current_parent not in self.data:
                    self.data[current_parent] = {}

            # Check for Child (Column 2)
            if row[1]:
                current_child = row[1]
                if current_parent and current_child not in self.data[current_parent]:
                    self.data[current_parent][current_child] = {}

            # Check for Subchild (Column 3)
            if row[2]:
                current_subchild = row[2]
                if (
                    current_parent
                    and current_child
                    and current_subchild not in self.data[current_parent][current_child]
                ):
                    self.data[current_parent][current_child][current_subchild] = {}

            if row[3] or row[4]:  # Ensure at least one of the columns has data
                key = row[3]
                value = row[4]

                # Ensure hierarchy exists for Parent, Child, Subchild
                if current_parent and current_child and current_subchild:
                    if key == "answer":  # Stop collecting options once we encounter "answer"
                        self.data[current_parent][current_child][current_subchild]["answer"] = value
                        collecting_options = False  # Stop collecting options

                    elif key == "options":  # Start collecting options
                        # Initialize options as an empty list if not already initialized
                        if "options" not in self.data[current_parent][current_child][current_subchild]:
                            self.data[current_parent][current_child][current_subchild]["options"] = []

                        # Append the value in the current row as the first option
                        if value:
                            self.data[current_parent][current_child][current_subchild]["options"].append(str(value))
                        
                        # Enable collecting options for subsequent rows
                        collecting_options = True

                    elif collecting_options:  # Collect options if in collection mode
                        if value:
                            self.data[current_parent][current_child][current_subchild]["options"].append(str(value))

                    else:  # For other keys, store them directly
                        self.data[current_parent][current_child][current_subchild][key] = value

    def save_to_json(self):
        """Saves the data as a JSON file."""
        with open(self.json_file, 'w') as json_file:
            json.dump(self.data, json_file, indent=4)

    def convert(self):
        """Full conversion process."""
        self.read_excel()
        self.save_to_json()

if __name__ == "__main__":
    # File paths
    excel_file = "/home/maheshreddy/Desktop/exceltojsonheirarchystructure/out.xlsx"
    json_file = "/home/maheshreddy/Desktop/exceltojsonheirarchystructure/example5.json"

    # Create the converter and process the Excel file
    converter = ExcelToJson(excel_file, json_file)
    converter.convert()

    print(f"JSON file has been created at: {json_file}")


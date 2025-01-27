"""TableSpam provides a formla-based syntax to define good-enough tables."""

from tablespam._Formula.Formulas import Formula
from tablespam._as_string.as_string import tbl_as_string
import polars as pl
import great_tables as gt
import openpyxl as opy
from tablespam.GT._as_gt.as_gt import (
    add_gt_spanners,
    add_gt_rowname_separator,
    add_gt_titles,
    add_gt_footnote,
    FormattingFunction,
)
from tablespam.GT.formatting import default_formatting
from tablespam.Excel.xlsx_styles import XlsxStyles
from tablespam.Excel._as_excel.as_excel import tbl_as_excel


class TableSpam:
    """Create complex table spanners with a simple formula.

    `tablespam` provides a formula-based approach to adding headers and spanners
    to an existing polars DataFrame. The goal is to provide a unified, easy-to-use,
    but good-enough approach to building and exporting tables to Excel, HTML,
    and LaTeX. To this end, `tablespam` leverages the powerful packages `openpyxl`
    and `great_tables`.

    The table headers are defined with a basic formula-inspired approach. For
    example, `Species ~ Sepal_Length + Sepal_Width` defines a table with `Species`
    as the row labels and `Sepal_Length` and `Sepal_Width` as columns. The output
    will look like:

    .. code-block::

        | Species | Sepal_Length | Sepal_Width |
        |:--------|-------------:|------------:|
        | setosa  |          5.1 |         3.5 |
        | setosa  |          4.9 |         3.0 |


    Note that the row labels (`Species`) are in a separate block to the left.

    You can add spanner labels as follows:

    `Species ~ (Sepal = Sepal_Length + Sepal_Width) + (Petal = Petal_Length + Petal_Width)`

    This results in an output like:

    .. code-block::

        |         |      Sepal     |      Petal     |
        | Species | Length | Width | Length | Width |
        |:--------|-------:|------:|-------:|------:|
        | setosa  |    5.1 |   3.5 |    1.4 |   0.2 |


    Nested spanners can also be defined, e.g.,
    `Species ~ (Sepal = (Length = Sepal_Length) + (Width = Sepal_Width))`.

    When exporting tables, you can rename columns in the headers. For example,
    `Species ~ (Sepal = Length:Sepal_Length + Width:Sepal_Width)` creates:

    .. code-block::

        |         |      Sepal     |      Petal     |
        | Species | Length | Width | Length | Width |
        |:--------|-------:|------:|-------:|------:|
        | setosa  |    5.1 |   3.5 |    1.4 |   0.2 |

    To create a table without row labels, use:

    `1 ~ (Sepal = Length:Sepal_Length + Width:Sepal_Width) + (Petal = Length:Petal_Length + Width:Petal_Width)`

    This creates:

    .. code-block::

        |      Sepal     |      Petal     |
        | Length | Width | Length | Width |
        |-------:|------:|-------:|------:|
        |    5.1 |   3.5 |    1.4 |   0.2 |

    Tables created with `tablespam` can be exported to Excel (using `openpyxl`),
    HTML (using `great_tables`), and LaTeX (using `great_tables`).

    References:
    - `openpyxl`: https://openpyxl.readthedocs.io/
    - `great_tables`: https://posit-dev.github.io/great-tables/articles/intro.html

    Examples:
        >>> from tablespam import TableSpam
        >>> import polars as pl
        >>> cars = pl.DataFrame(
        ...     {
        ...         'mpg': [21.0, 21.0, 22.8, 21.4, 18.7, 18.1, 14.3, 24.4, 22.8, 19.2],
        ...         'cyl': [6, 6, 4, 6, 8, 6, 8, 4, 4, 6],
        ...         'disp': [160.0, 160.0, 108.0, 258.0, 360.0, 225.0, 360.0, 146.7, 140.8, 167.6,],
        ...         'hp': [110, 110, 93, 110, 175, 105, 245, 62, 95, 123],
        ...         'drat': [3.90, 3.90, 3.85, 3.08, 3.15, 2.76, 3.21, 4.08, 3.92, 3.92,],
        ...         'wt': [2.620, 2.875, 2.320, 3.215, 3.440, 3.460, 3.570, 3.190, 3.150, 3.440,],
        ...         'qsec': [16.46, 17.02, 18.61, 19.44, 17.02, 20.22, 15.84, 20.00, 22.90, 18.30,],
        ...         'vs': [0, 0, 1, 1, 0, 1, 0, 1, 1, 1],
        ...         'am': [1, 1, 1, 0, 0, 0, 0, 0, 0, 0],
        ...         'gear': [4, 4, 4, 3, 3, 3, 3, 4, 4, 4],
        ...         'carb': [4, 4, 1, 1, 2, 1, 4, 2, 2, 4],
        ...     }
        ... )
        >>> summarized_table = (
        ...     cars.group_by(['cyl', 'vs'])
        ...     .agg(
        ...         [
        ...             pl.len().alias('N'),
        ...             pl.col('hp').mean().alias('mean_hp'),
        ...             pl.col('hp').std().alias('sd_hp'),
        ...             pl.col('wt').mean().alias('mean_wt'),
        ...             pl.col('wt').std().alias('sd_wt'),
        ...         ]
        ...     )
        ...     .sort(['cyl', 'vs'])
        ... )
        >>> tbl = TableSpam(
        ...     data=summarized_table,
        ...     formula='''Cylinder:cyl + Engine:vs ~
        ...                 N +
        ...                 (`Horse Power` = Mean:mean_hp + SD:sd_hp) +
        ...                 (`Weight` = Mean:mean_wt + SD:sd_wt)''',
        ...     title='Motor Trend Car Road Tests',
        ...     subtitle='A table created with tablespam',
        ...     footnote='Data from the infamous mtcars data set.',
        ... )
        >>> print(tbl.as_string())
        Motor Trend Car Road Tests
        A table created with tablespam
        <BLANKLINE>
        |                 |     Horse Power      Weight      |
        | Cylinder Engine | N   Mean        SD   Mean   SD   |
        | -------- ------ - --- ----------- ---- ------ ---- |
        | 4        1      | 3   83.33       18.5 2.89   0.49 |
        | 6        0      | 2   110.0       0.0  2.75   0.18 |
        | 6        1      | 3   112.67      9.29 3.37   0.14 |
        | ...      ...    | ... ...         ...  ...    ...  |
        Data from the infamous mtcars data set.
        <BLANKLINE>
    """

    def __init__(
        self,
        data: pl.DataFrame,
        formula: str,
        title: str | None = None,
        subtitle: str | None = None,
        footnote: str | None = None,
    ):
        """Create complex table spanners with a simple formula.

        `tablespam` provides a formula-based approach to adding headers and spanners
        to an existing polars DataFrame. The goal is to provide a unified, easy-to-use,
        but good-enough approach to building and exporting tables to Excel, HTML,
        and LaTeX. To this end, `tablespam` leverages the powerful packages `openpyxl`
        and `great_tables`.

        The table headers are defined with a basic formula-inspired approach. For
        example, `Species ~ Sepal_Length + Sepal_Width` defines a table with `Species`
        as the row labels and `Sepal_Length` and `Sepal_Width` as columns. The output
        will look like:

        .. code-block::

            | Species | Sepal_Length | Sepal_Width |
            |:--------|-------------:|------------:|
            | setosa  |          5.1 |         3.5 |
            | setosa  |          4.9 |         3.0 |


        Note that the row labels (`Species`) are in a separate block to the left.

        You can add spanner labels as follows:

        `Species ~ (Sepal = Sepal_Length + Sepal_Width) + (Petal = Petal_Length + Petal_Width)`

        This results in an output like:

        .. code-block::

            |         |      Sepal     |      Petal     |
            | Species | Length | Width | Length | Width |
            |:--------|-------:|------:|-------:|------:|
            | setosa  |    5.1 |   3.5 |    1.4 |   0.2 |


        Nested spanners can also be defined, e.g.,
        `Species ~ (Sepal = (Length = Sepal_Length) + (Width = Sepal_Width))`.

        When exporting tables, you can rename columns in the headers. For example,
        `Species ~ (Sepal = Length:Sepal_Length + Width:Sepal_Width)` creates:

        .. code-block::

            |         |      Sepal     |      Petal     |
            | Species | Length | Width | Length | Width |
            |:--------|-------:|------:|-------:|------:|
            | setosa  |    5.1 |   3.5 |    1.4 |   0.2 |

        To create a table without row labels, use:

        `1 ~ (Sepal = Length:Sepal_Length + Width:Sepal_Width) + (Petal = Length:Petal_Length + Width:Petal_Width)`

        This creates:

        .. code-block::

            |      Sepal     |      Petal     |
            | Length | Width | Length | Width |
            |-------:|------:|-------:|------:|
            |    5.1 |   3.5 |    1.4 |   0.2 |

        Tables created with `tablespam` can be exported to Excel (using `openpyxl`),
        HTML (using `great_tables`), and LaTeX (using `great_tables`).

        References:
        - `openpyxl`: https://openpyxl.readthedocs.io/
        - `great_tables`: https://posit-dev.github.io/great-tables/articles/intro.html

        Args:
            data (pl.DataFrame): Polars data frame with the data that should be shown in the table.
            formula (str): The tables for TableSpam are described in a single formula. See above for a detailed description.
            title (str | None, optional): The title of the table. Defaults to None.
            subtitle (str | None, optional): The subtitle of the table. Defaults to None.
            footnote (str | None, optional): The footnote of the table. Defaults to None.

        Returns:
            TableSpam: An object containing the title, subtitle, header info, data, and footnote.

        Examples:
            >>> from tablespam import TableSpam
            >>> import polars as pl
            >>> cars = pl.DataFrame(
            ...     {
            ...         'mpg': [21.0, 21.0, 22.8, 21.4, 18.7, 18.1, 14.3, 24.4, 22.8, 19.2],
            ...         'cyl': [6, 6, 4, 6, 8, 6, 8, 4, 4, 6],
            ...         'disp': [160.0, 160.0, 108.0, 258.0, 360.0, 225.0, 360.0, 146.7, 140.8, 167.6,],
            ...         'hp': [110, 110, 93, 110, 175, 105, 245, 62, 95, 123],
            ...         'drat': [3.90, 3.90, 3.85, 3.08, 3.15, 2.76, 3.21, 4.08, 3.92, 3.92,],
            ...         'wt': [2.620, 2.875, 2.320, 3.215, 3.440, 3.460, 3.570, 3.190, 3.150, 3.440,],
            ...         'qsec': [16.46, 17.02, 18.61, 19.44, 17.02, 20.22, 15.84, 20.00, 22.90, 18.30,],
            ...         'vs': [0, 0, 1, 1, 0, 1, 0, 1, 1, 1],
            ...         'am': [1, 1, 1, 0, 0, 0, 0, 0, 0, 0],
            ...         'gear': [4, 4, 4, 3, 3, 3, 3, 4, 4, 4],
            ...         'carb': [4, 4, 1, 1, 2, 1, 4, 2, 2, 4],
            ...     }
            ... )
            >>> summarized_table = (
            ...     cars.group_by(['cyl', 'vs'])
            ...     .agg(
            ...         [
            ...             pl.len().alias('N'),
            ...             pl.col('hp').mean().alias('mean_hp'),
            ...             pl.col('hp').std().alias('sd_hp'),
            ...             pl.col('wt').mean().alias('mean_wt'),
            ...             pl.col('wt').std().alias('sd_wt'),
            ...         ]
            ...     )
            ...     .sort(['cyl', 'vs'])
            ... )
            >>> tbl = TableSpam(
            ...     data=summarized_table,
            ...     formula='''Cylinder:cyl + Engine:vs ~
            ...                 N +
            ...                 (`Horse Power` = Mean:mean_hp + SD:sd_hp) +
            ...                 (`Weight` = Mean:mean_wt + SD:sd_wt)''',
            ...     title='Motor Trend Car Road Tests',
            ...     subtitle='A table created with tablespam',
            ...     footnote='Data from the infamous mtcars data set.',
            ... )
            >>> print(tbl.as_string())
            Motor Trend Car Road Tests
            A table created with tablespam
            <BLANKLINE>
            |                 |     Horse Power      Weight      |
            | Cylinder Engine | N   Mean        SD   Mean   SD   |
            | -------- ------ - --- ----------- ---- ------ ---- |
            | 4        1      | 3   83.33       18.5 2.89   0.49 |
            | 6        0      | 2   110.0       0.0  2.75   0.18 |
            | 6        1      | 3   112.67      9.29 3.37   0.14 |
            | ...      ...    | ... ...         ...  ...    ...  |
            Data from the infamous mtcars data set.
            <BLANKLINE>
        """
        self.data = data

        form = Formula(formula=formula)
        variables = form.get_variables()
        self.table_data = {
            'row_data': select_data(self.data, variables['lhs']),
            'col_data': select_data(self.data, variables['rhs']),
        }

        self.title = title
        self.subtitle = subtitle
        self.footnote = footnote
        self.header = form.get_entries()

    def __repr__(self) -> str:
        """Print the TableSpam table.

        Returns:
            str: String representing table
        """
        return tbl_as_string(self)

    def as_string(self, digits: int = 2, n: int = 3, max_char: int = 30) -> str:
        """Translates a table to string.

        The main purpose if this transformation is for debugging. Exporting to gt or excel
        are more feature complete.

        Args:
            digits (int, optional): Number of digits to round floats to. Defaults to 2.
            n (int, optional): number of rows from the data set to print. Defaults to 3.
            max_char (int, optional): number of characters that each cell at maximum is allows to have. Defaults to 30.

        Returns:
            str: String describing the table

        Examples:
            >>> from tablespam import TableSpam
            >>> import polars as pl
            >>> cars = pl.DataFrame(
            ...     {
            ...         'mpg': [21.0, 21.0, 22.8, 21.4, 18.7, 18.1, 14.3, 24.4, 22.8, 19.2],
            ...         'cyl': [6, 6, 4, 6, 8, 6, 8, 4, 4, 6],
            ...         'disp': [160.0, 160.0, 108.0, 258.0, 360.0, 225.0, 360.0, 146.7, 140.8, 167.6,],
            ...         'hp': [110, 110, 93, 110, 175, 105, 245, 62, 95, 123],
            ...         'drat': [3.90, 3.90, 3.85, 3.08, 3.15, 2.76, 3.21, 4.08, 3.92, 3.92,],
            ...         'wt': [2.620, 2.875, 2.320, 3.215, 3.440, 3.460, 3.570, 3.190, 3.150, 3.440,],
            ...         'qsec': [16.46, 17.02, 18.61, 19.44, 17.02, 20.22, 15.84, 20.00, 22.90, 18.30,],
            ...         'vs': [0, 0, 1, 1, 0, 1, 0, 1, 1, 1],
            ...         'am': [1, 1, 1, 0, 0, 0, 0, 0, 0, 0],
            ...         'gear': [4, 4, 4, 3, 3, 3, 3, 4, 4, 4],
            ...         'carb': [4, 4, 1, 1, 2, 1, 4, 2, 2, 4],
            ...     }
            ... )
            >>> summarized_table = (
            ...     cars.group_by(['cyl', 'vs'])
            ...     .agg(
            ...         [
            ...             pl.len().alias('N'),
            ...             pl.col('hp').mean().alias('mean_hp'),
            ...             pl.col('hp').std().alias('sd_hp'),
            ...             pl.col('wt').mean().alias('mean_wt'),
            ...             pl.col('wt').std().alias('sd_wt'),
            ...         ]
            ...     )
            ...     .sort(['cyl', 'vs'])
            ... )
            >>> tbl = TableSpam(
            ...     data=summarized_table,
            ...     formula='''Cylinder:cyl + Engine:vs ~
            ...                 N +
            ...                 (`Horse Power` = Mean:mean_hp + SD:sd_hp) +
            ...                 (`Weight` = Mean:mean_wt + SD:sd_wt)''',
            ...     title='Motor Trend Car Road Tests',
            ...     subtitle='A table created with tablespam',
            ...     footnote='Data from the infamous mtcars data set.',
            ... )
            >>> print(tbl.as_string())
            Motor Trend Car Road Tests
            A table created with tablespam
            <BLANKLINE>
            |                 |     Horse Power      Weight      |
            | Cylinder Engine | N   Mean        SD   Mean   SD   |
            | -------- ------ - --- ----------- ---- ------ ---- |
            | 4        1      | 3   83.33       18.5 2.89   0.49 |
            | 6        0      | 2   110.0       0.0  2.75   0.18 |
            | 6        1      | 3   112.67      9.29 3.37   0.14 |
            | ...      ...    | ... ...         ...  ...    ...  |
            Data from the infamous mtcars data set.
            <BLANKLINE>
        """
        return tbl_as_string(self, digits=digits, n=n, max_char=max_char)

    def as_gt(
        self,
        separator_style: gt.style.borders = gt.style.borders(
            sides=['right'], color='gray'
        ),
        formatting: FormattingFunction | None = default_formatting,
        groupname_col: str | None = None,
        auto_align: bool = True,
        id: str | None = None,
        locale: str | None = None,
    ) -> gt.GT:
        """Translates a table created with `tablespam` into a `gt` table.

        The `tablespam` library does not provide built-in support for rendering tables as HTML.
        However, with `as_gt`, a `tablespam` table can be converted into a `gt` table,
        which supports HTML and LaTeX output. For more details on `gt`, see
        <https://gt.rstudio.com/>.

        Args:
            groupname_col (str, optional): Column names to group data. Refer to the
                `gt` documentation for details.
            separator_style (str, optional): Style of the vertical line separating row
                names from data.
            formatting (function, optional): This function is applied to the gt to format
                all columns.
            auto_align (bool, optional): Should the table entries be aligned automatically? See great_tables for more information
            id (str, optional): Id of the HTML table. See great_tables for more details
            locale (str, optional): affects formatting of dates and numbers. See great_tables for more details.

        Returns:
            GtTable: A `gt` table object that can be further customized using the `gt` package.

        Examples:
            >>> from tablespam import TableSpam
            >>> import polars as pl
            >>> cars = pl.DataFrame(
            ...     {
            ...         'mpg': [21.0, 21.0, 22.8, 21.4, 18.7, 18.1, 14.3, 24.4, 22.8, 19.2],
            ...         'cyl': [6, 6, 4, 6, 8, 6, 8, 4, 4, 6],
            ...         'disp': [160.0, 160.0, 108.0, 258.0, 360.0, 225.0, 360.0, 146.7, 140.8, 167.6,],
            ...         'hp': [110, 110, 93, 110, 175, 105, 245, 62, 95, 123],
            ...         'drat': [3.90, 3.90, 3.85, 3.08, 3.15, 2.76, 3.21, 4.08, 3.92, 3.92,],
            ...         'wt': [2.620, 2.875, 2.320, 3.215, 3.440, 3.460, 3.570, 3.190, 3.150, 3.440,],
            ...         'qsec': [16.46, 17.02, 18.61, 19.44, 17.02, 20.22, 15.84, 20.00, 22.90, 18.30,],
            ...         'vs': [0, 0, 1, 1, 0, 1, 0, 1, 1, 1],
            ...         'am': [1, 1, 1, 0, 0, 0, 0, 0, 0, 0],
            ...         'gear': [4, 4, 4, 3, 3, 3, 3, 4, 4, 4],
            ...         'carb': [4, 4, 1, 1, 2, 1, 4, 2, 2, 4],
            ...     }
            ... )
            >>> summarized_table = (
            ...     cars.group_by(['cyl', 'vs'])
            ...     .agg(
            ...         [
            ...             pl.len().alias('N'),
            ...             pl.col('hp').mean().alias('mean_hp'),
            ...             pl.col('hp').std().alias('sd_hp'),
            ...             pl.col('wt').mean().alias('mean_wt'),
            ...             pl.col('wt').std().alias('sd_wt'),
            ...         ]
            ...     )
            ...     .sort(['cyl', 'vs'])
            ... )
            >>> tbl = TableSpam(
            ...     data=summarized_table,
            ...     formula='''Cylinder:cyl + Engine:vs ~
            ...                 N +
            ...                 (`Horse Power` = Mean:mean_hp + SD:sd_hp) +
            ...                 (`Weight` = Mean:mean_wt + SD:sd_wt)''',
            ...     title='Motor Trend Car Road Tests',
            ...     subtitle='A table created with tablespam',
            ...     footnote='Data from the infamous mtcars data set.',
            ... )
            >>> gt_tbl = tbl.as_gt()
            >>> # Use tbl.as_gt().show() to show the table in the browser.
        """
        if (
            (self.header['lhs'] is not None)
            and (self.table_data['row_data'] is not None)
            and (isinstance(self.table_data['row_data'], pl.DataFrame))
            and (isinstance(self.table_data['col_data'], pl.DataFrame))
        ):
            data_set = pl.concat(
                [self.table_data['row_data'], self.table_data['col_data']],
                how='horizontal',
            )
        elif isinstance(self.table_data['col_data'], pl.DataFrame):
            data_set = self.table_data['col_data']
        else:
            raise ValueError('table_data should be of type pl.DataFrame.')

        # Create the gt-like table (assuming `gt` functionality is implemented)
        gt_tbl = gt.GT(
            data=data_set,
            groupname_col=groupname_col,
            auto_align=auto_align,
            id=id,
            locale=locale,
        )

        gt_tbl = add_gt_spanners(gt_tbl=gt_tbl, tbl=self)

        if (self.header['lhs'] is not None) and (
            self.table_data['row_data'] is not None
        ):
            rowname_headers = self.table_data['row_data'].columns
            gt_tbl = add_gt_rowname_separator(
                gt_tbl=gt_tbl,
                right_of=rowname_headers[-1],  # Use the last row header
                separator_style=separator_style,
            )

        # Add titles and subtitles if present
        if self.title is not None or self.subtitle is not None:
            gt_tbl = add_gt_titles(
                gt_tbl=gt_tbl, title=self.title, subtitle=self.subtitle
            )

        # Add footnotes if present
        if self.footnote is not None:
            gt_tbl = add_gt_footnote(gt_tbl=gt_tbl, footnote=self.footnote)

        # Apply auto-formatting if requested
        if formatting is not None:
            gt_tbl = default_formatting(gt_tbl)

        return gt_tbl

    def as_excel(
        self,
        workbook: opy.Workbook | None = None,
        sheet: str = 'Table',
        start_row: int = 1,
        start_col: int = 1,
        styles: XlsxStyles | None = None,
    ) -> opy.Workbook:
        """Export a TableSpam table to Excel.

        Tablespam uses openpyxl to export tables to Excel workbooks. See
        https://openpyxl.readthedocs.io/en/stable/ for more details on openpyxl.

        Args:
            workbook (opy.Workbook | None, optional): An openpyxl workbook to which the table should be added.
                When set to None, a new workbook will be created. Defaults to None.
            sheet (str, optional): The name of the sheet to which the table should be written. Defaults to 'Table'.
            start_row (int, optional): Index of the row where the table starts in the sheet. Defaults to 1.
            start_col (int, optional): Index of the column where the table starts in the sheet. Defaults to 1.
            styles (XlsxStyles | None, optional): Custom styles that are applied to the table. Defaults to None.

        Returns:
            opy.Workbook: openpyxl workbook


        Examples:
            >>> from tablespam import TableSpam
            >>> import polars as pl
            >>> cars = pl.DataFrame(
            ...     {
            ...         'mpg': [21.0, 21.0, 22.8, 21.4, 18.7, 18.1, 14.3, 24.4, 22.8, 19.2],
            ...         'cyl': [6, 6, 4, 6, 8, 6, 8, 4, 4, 6],
            ...         'disp': [160.0, 160.0, 108.0, 258.0, 360.0, 225.0, 360.0, 146.7, 140.8, 167.6,],
            ...         'hp': [110, 110, 93, 110, 175, 105, 245, 62, 95, 123],
            ...         'drat': [3.90, 3.90, 3.85, 3.08, 3.15, 2.76, 3.21, 4.08, 3.92, 3.92,],
            ...         'wt': [2.620, 2.875, 2.320, 3.215, 3.440, 3.460, 3.570, 3.190, 3.150, 3.440,],
            ...         'qsec': [16.46, 17.02, 18.61, 19.44, 17.02, 20.22, 15.84, 20.00, 22.90, 18.30,],
            ...         'vs': [0, 0, 1, 1, 0, 1, 0, 1, 1, 1],
            ...         'am': [1, 1, 1, 0, 0, 0, 0, 0, 0, 0],
            ...         'gear': [4, 4, 4, 3, 3, 3, 3, 4, 4, 4],
            ...         'carb': [4, 4, 1, 1, 2, 1, 4, 2, 2, 4],
            ...     }
            ... )
            >>> summarized_table = (
            ...     cars.group_by(['cyl', 'vs'])
            ...     .agg(
            ...         [
            ...             pl.len().alias('N'),
            ...             pl.col('hp').mean().alias('mean_hp'),
            ...             pl.col('hp').std().alias('sd_hp'),
            ...             pl.col('wt').mean().alias('mean_wt'),
            ...             pl.col('wt').std().alias('sd_wt'),
            ...         ]
            ...     )
            ...     .sort(['cyl', 'vs'])
            ... )
            >>> tbl = TableSpam(
            ...     data=summarized_table,
            ...     formula='''Cylinder:cyl + Engine:vs ~
            ...                 N +
            ...                 (`Horse Power` = Mean:mean_hp + SD:sd_hp) +
            ...                 (`Weight` = Mean:mean_wt + SD:sd_wt)''',
            ...     title='Motor Trend Car Road Tests',
            ...     subtitle='A table created with tablespam',
            ...     footnote='Data from the infamous mtcars data set.',
            ... )
            >>> wb = tbl.as_excel()  # Export to Excel workbook
            >>> # wb.save("tablespam_table.xlsx") # Write to an Excel file.
        """
        if workbook is None:
            workbook = opy.Workbook()
            # openpyxl automatically adds a default sheet
            # that we will remove
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
        if styles is None:
            styles = XlsxStyles()
        if sheet not in workbook.sheetnames:
            workbook.create_sheet(title=sheet)

        wb = tbl_as_excel(
            tbl=self,
            workbook=workbook,
            sheet=sheet,
            start_row=start_row,
            start_col=start_col,
            styles=styles,
        )
        return wb


def select_data(data: pl.DataFrame, variables: list[str]) -> pl.DataFrame | None:
    """Subsets the data frame to only the relevant variables.

    Args:
        data (pl.DataFrame): polars data frame that should be subsetted
        variables (list[str]): list with names of items that should be retained

    Returns:
        pl.DataFrame | None: polars data frame with the specified variables
    """
    if variables is None:
        return None
    return data.select(variables)

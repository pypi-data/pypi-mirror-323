"""Export a TableSpam table to Excel."""

from __future__ import annotations
from typing import TYPE_CHECKING, Callable, Any, cast

import openpyxl as opy
from openpyxl.utils import get_column_interval
from openpyxl.cell.cell import Cell
import polars as pl
from tablespam.Excel._as_excel.write_excel import write_excel_col
from tablespam.Excel.xlsx_styles import XlsxStyles
from tablespam.Excel._as_excel.styles import set_region_style
from tablespam.Excel._as_excel.locations import Locations
import numpy as np

if TYPE_CHECKING:
    from tablespam.TableSpam import TableSpam
    from tablespam._Formula.Entry import HeaderEntry


def tbl_as_excel(
    tbl: TableSpam,
    workbook: opy.Workbook,
    sheet: str = 'Table',
    start_row: int = 1,
    start_col: int = 1,
    styles: XlsxStyles | None = None,
) -> opy.Workbook:
    """Export a TableSpam table to Excel.

    Args:
        tbl (TableSpam): TableSpam table created with TableSpam
        workbook (opy.Workbook): openpyxl workbook
        sheet (str, optional): name of the sheet to which the table should be added. Defaults to 'Table'.
        start_row (int, optional): index of the row at which the table should start. Defaults to 1.
        start_col (int, optional): index of the column at which the table should start. Defaults to 1.
        styles (XlsxStyles | None, optional): Styles that should be applied to the table. Defaults to None.

    Returns:
        opy.Workbook: workbook with added table
    """
    if styles is None:
        styles = XlsxStyles()

    locations = Locations(tbl=tbl, start_row=start_row, start_col=start_col)

    fill_background(
        tbl=tbl, workbook=workbook, sheet=sheet, locations=locations, styles=styles
    )

    write_title(
        tbl=tbl, workbook=workbook, sheet=sheet, locations=locations, styles=styles
    )

    write_header(
        workbook=workbook,
        sheet=sheet,
        header=tbl.header,
        locations=locations,
        styles=styles,
    )

    write_data(
        workbook=workbook,
        sheet=sheet,
        header=tbl.header,
        table_data=tbl.table_data,
        locations=locations,
        styles=styles,
    )

    write_footnote(
        tbl=tbl, workbook=workbook, sheet=sheet, locations=locations, styles=styles
    )

    # We create the outlines last as we may have to overwrite some border colors.
    create_outlines(
        tbl=tbl, workbook=workbook, sheet=sheet, locations=locations, styles=styles
    )

    return workbook


def fill_background(
    tbl: TableSpam,
    workbook: opy.Workbook,
    sheet: str,
    locations: Locations,
    styles: XlsxStyles,
) -> None:
    """Fill the background of the Excel table.

    Args:
        tbl (TableSpam): TableSpam table created with TableSpam
        workbook (opy.Workbook): openpyxl workbook
        sheet (str, optional): name of the sheet to which the table should be added. Defaults to 'Table'.
        locations (dict[str, dict[str, int  |  None]]): a dict describing the locations (indexes) of different elements found in the table.
        styles (XlsxStyles): Styles that should be applied to the table.
    """
    sheet_ref = workbook[sheet]

    # Title
    if tbl.title is not None:
        set_region_style(
            sheet=sheet_ref,
            style=styles.bg_title,
            start_col=locations.get_col('start_col_title'),
            end_col=locations.get_col('end_col_title'),
            start_row=locations.get_row('start_row_title'),
            end_row=locations.get_row('end_row_title'),
        )

    # Subtitle
    if tbl.subtitle is not None:
        set_region_style(
            sheet=sheet_ref,
            style=styles.bg_subtitle,
            start_col=locations.get_col('start_col_subtitle'),
            end_col=locations.get_col('end_col_subtitle'),
            start_row=locations.get_row('start_row_subtitle'),
            end_row=locations.get_row('end_row_subtitle'),
        )

    # Header LHS
    if tbl.header['lhs'] is not None:
        set_region_style(
            sheet=sheet_ref,
            style=styles.bg_header_lhs,
            start_col=locations.get_col('start_col_header_lhs'),
            end_col=locations.get_col('end_col_header_lhs'),
            start_row=locations.get_row('start_row_header'),
            end_row=locations.get_row('end_row_header'),
        )

    # Header RHS
    set_region_style(
        sheet=sheet_ref,
        style=styles.bg_header_rhs,
        start_col=locations.get_col('start_col_header_rhs'),
        end_col=locations.get_col('end_col_header_rhs'),
        start_row=locations.get_row('start_row_header'),
        end_row=locations.get_row('end_row_header'),
    )

    # Rownames
    if tbl.header['lhs'] is not None:
        set_region_style(
            sheet=sheet_ref,
            style=styles.bg_rownames,
            start_col=locations.get_col('start_col_header_lhs'),
            end_col=locations.get_col('end_col_header_lhs'),
            start_row=locations.get_row('start_row_data'),
            end_row=locations.get_row('end_row_data'),
        )

    # Data
    set_region_style(
        sheet=sheet_ref,
        style=styles.bg_data,
        start_col=locations.get_col('start_col_header_rhs'),
        end_col=locations.get_col('end_col_header_rhs'),
        start_row=locations.get_row('start_row_data'),
        end_row=locations.get_row('end_row_data'),
    )

    # Footnote
    if tbl.footnote is not None:
        set_region_style(
            sheet=sheet_ref,
            style=styles.bg_footnote,
            start_col=locations.get_col('start_col_footnote'),
            end_col=locations.get_col('end_col_footnote'),
            start_row=locations.get_row('start_row_footnote'),
            end_row=locations.get_row('end_row_footnote'),
        )


def write_title(
    tbl: TableSpam,
    workbook: opy.Workbook,
    sheet: str,
    locations: Locations,
    styles: XlsxStyles,
) -> None:
    """Write the title and subtitle to the Excel workbook.

    Args:
        tbl (TableSpam): TableSpam table created with TableSpam
        workbook (opy.Workbook): openpyxl workbook
        sheet (str, optional): name of the sheet to which the table should be added. Defaults to 'Table'.
        locations (dict[str, dict[str, int  |  None]]): a dict describing the locations (indexes) of different elements found in the table.
        styles (XlsxStyles): Styles that should be applied to the table.
    """
    if tbl.title is not None:
        loc = get_column_interval(
            start=locations.get_col('start_col_title'),
            end=locations.get_col('start_col_title'),
        )[0] + str(locations.get_row('start_row_title'))
        workbook[sheet][loc] = tbl.title

        workbook[sheet].merge_cells(
            start_row=locations.get_row('start_row_title'),
            start_column=locations.get_col('start_col_title'),
            end_row=locations.get_row('start_row_title'),
            end_column=locations.get_col('end_col_title'),
        )
        set_region_style(
            sheet=workbook[sheet],
            style=styles.cell_title,
            start_row=locations.get_row('start_row_title'),
            start_col=locations.get_col('start_col_title'),
            end_row=locations.get_row('start_row_title'),
            end_col=locations.get_col('end_col_title'),
        )

    if tbl.subtitle is not None:
        loc = get_column_interval(
            start=locations.get_col('start_col_subtitle'),
            end=locations.get_col('start_col_subtitle'),
        )[0] + str(locations.get_row('start_row_subtitle'))
        workbook[sheet][loc] = tbl.subtitle
        workbook[sheet].merge_cells(
            start_row=locations.get_row('start_row_subtitle'),
            start_column=locations.get_col('start_col_subtitle'),
            end_row=locations.get_row('start_row_subtitle'),
            end_column=locations.get_col('end_col_subtitle'),
        )
        set_region_style(
            sheet=workbook[sheet],
            style=styles.cell_subtitle,
            start_row=locations.get_row('start_row_subtitle'),
            start_col=locations.get_col('start_col_subtitle'),
            end_row=locations.get_row('start_row_subtitle'),
            end_col=locations.get_col('end_col_subtitle'),
        )


def write_header(
    workbook: opy.Workbook,
    sheet: str,
    header: dict[str, HeaderEntry],
    locations: Locations,
    styles: XlsxStyles,
) -> None:
    """Fill the background of the Excel table.

    Args:
        workbook (opy.Workbook): openpyxl workbook
        sheet (str): name of the sheet to which the table should be added. Defaults to 'Table'.
        header (dict[str, HeaderEntry]): header information for right and left hand side of the table.
        locations (dict[str, dict[str, int  |  None]]): a dict describing the locations (indexes) of different elements found in the table.
        styles (XlsxStyles): Styles that should be applied to the table.
    """
    if header['lhs'] is not None:
        max_level = max(header['lhs'].level, header['rhs'].level)

        write_header_entry(
            workbook=workbook,
            sheet=sheet,
            header_entry=header['lhs'],
            max_level=max_level,
            start_row=locations.get_row('start_row_header'),
            start_col=locations.get_col('start_col_header_lhs'),
            style=styles.cell_header_lhs,
        )
    else:
        max_level = header['rhs'].level

    write_header_entry(
        workbook=workbook,
        sheet=sheet,
        header_entry=header['rhs'],
        max_level=max_level,
        start_row=locations.get_row('start_row_header'),
        start_col=locations.get_col('start_col_header_rhs'),
        style=styles.cell_header_rhs,
    )


def write_header_entry(
    workbook: opy.Workbook,
    sheet: str,
    header_entry: HeaderEntry,
    max_level: int,
    start_row: int,
    start_col: int,
    style: Callable[[Cell], None],
) -> None:
    """Add a specific header entry to the Excel workbook.

    Args:
        workbook (opy.Workbook): openpyxl workbook
        sheet (str): name of the sheet to which the table should be added. Defaults to 'Table'.
        header_entry (HeaderEntry): specific header entry that will be added to the workbook.
        max_level (int): The highest level of the header entries.
        start_row (int): At what row should the current header entry be added?
        start_col (int): At what column should the current header entry be added?
        style (Callable[[Cell], None]): style to be added to the entry.
    """
    # write current entry name into table
    if header_entry.name != '_BASE_LEVEL_':
        loc = get_column_interval(start=start_col, end=start_col)[0] + str(
            start_row + (max_level - header_entry.level) - 1
        )
        workbook[sheet][loc] = header_entry.name

        if header_entry.width > 1:
            workbook[sheet].merge_cells(
                start_row=start_row + (max_level - header_entry.level) - 1,
                start_column=start_col,
                end_row=start_row + (max_level - header_entry.level) - 1,
                end_column=start_col + header_entry.width - 1,
            )

        set_region_style(
            sheet=workbook[sheet],
            style=style,
            start_row=start_row + (max_level - header_entry.level) - 1,
            start_col=start_col,
            end_row=start_row + (max_level - header_entry.level) - 1,
            end_col=start_col + header_entry.width - 1,
        )

    # entries may have sub-entries, that also have to be written down
    start_col_entry = start_col
    for entry in header_entry.entries:
        write_header_entry(
            workbook=workbook,
            sheet=sheet,
            header_entry=entry,
            max_level=max_level,
            start_row=start_row,
            start_col=start_col_entry,
            style=style,
        )
        start_col_entry += entry.width


def row_data_cell_ids(row_data: pl.DataFrame) -> np.ndarray[Any, Any]:
    """Generate unique IDs to represent entries that should be merged.

    Args:
        row_data (pl.DataFrame): data that is written as rownames in the table.

    Returns:
        np.ndarray[Any]: a matrix with the same number of rows and columns as the row_data. Each entry is given an index. If two cells should be merged, they will have the same index.
    """
    ids = np.full(
        (row_data.shape[0], row_data.shape[1]),
        0,
        dtype=np.uint8,
    )
    ids[0, :] = 1
    if row_data.shape[0] == 1:
        return ids

    for ro in range(1, row_data.shape[0]):
        for co in range(0, row_data.shape[1]):
            if row_data[ro, range(0, co + 1)].equals(
                row_data[ro - 1, range(0, co + 1)]
            ):
                ids[ro, co] = ids[ro - 1, co]
            else:
                ids[ro, co] = ids[ro - 1, co] + 1
    return ids


def merge_rownames(
    workbook: opy.Workbook,
    sheet: str,
    table_data: dict[str, pl.DataFrame | None],
    locations: Locations,
    styles: XlsxStyles,
) -> None:
    """Merges consecutive rownames that are identical into a common cell.

    Args:
        workbook (opy.Workbook): openpyxl workbook
        sheet (str): name of the sheet to which the table should be added. Defaults to 'Table'.
        table_data (dict[str, pl.DataFrame]): Data that should be written into the table body.
        locations (dict[str, dict[str, int  |  None]]): a dict describing the locations (indexes) of different elements found in the table.
        styles (XlsxStyles): Styles that should be applied to the table.

    Raises:
        ValueError: Error if data frame is None
    """
    if table_data['row_data'] is None:
        raise ValueError("table_data['row_data'] is None.")
    cell_ids = row_data_cell_ids(table_data['row_data'])

    # We merge all cells within a column that have the same id.
    for co in range(0, table_data['row_data'].shape[1]):
        unique_ids = set(cell_ids[:, co])
        for id in unique_ids:
            is_identical = [x == id for x in cell_ids[:, co]]
            if sum(is_identical) > 1:
                set_region_style(
                    sheet=workbook[sheet],
                    style=styles.merged_rownames_style,
                    start_row=locations.get_row('end_row_header')
                    + cast(int, np.min(np.where(is_identical)))
                    + 1,
                    start_col=locations.get_col('start_col_header_lhs') + cast(int, co),
                    end_row=locations.get_row('end_row_header')
                    + cast(int, np.max(np.where(is_identical)))
                    + 1,
                    end_col=locations.get_col('start_col_header_lhs') + cast(int, co),
                )
                workbook[sheet].merge_cells(
                    start_row=locations.get_row('end_row_header')
                    + cast(int, np.min(np.where(is_identical)))
                    + 1,
                    start_column=locations.get_col('start_col_header_lhs')
                    + cast(int, co),
                    end_row=locations.get_row('end_row_header')
                    + cast(int, np.max(np.where(is_identical)))
                    + 1,
                    end_column=locations.get_col('start_col_header_lhs')
                    + cast(int, co),
                )


def write_data(
    workbook: opy.Workbook,
    sheet: str,
    header: dict[str, HeaderEntry],
    table_data: dict[str, pl.DataFrame | None],
    locations: Locations,
    styles: XlsxStyles,
) -> None:
    """Write the data into the table body.

    Args:
        tbl (TableSpam): TableSpam table created with TableSpam
        workbook (opy.Workbook): openpyxl workbook
        sheet (str, optional): name of the sheet to which the table should be added. Defaults to 'Table'.
        header (dict[str, HeaderEntry]): header information for right and left hand side of the table.
        table_data (dict[str, pl.DataFrame]): data that should be written into the table body.
        locations (dict[str, dict[str, int  |  None]]): a dict describing the locations (indexes) of different elements found in the table.
        styles (XlsxStyles): Styles that should be applied to the table.

    Raises:
        ValueError: Error when row data does not exist.
        ValueError: Error when data does not exist.
        ValueError: Error when trying to add a style to a column that does not exist.
        ValueError: Error when trying to add a style to a row that does not exist.
    """
    if header['lhs'] is not None:
        if table_data['row_data'] is None:
            raise ValueError('Missing data')
        # Add row names and their styling
        i = 0
        for item in table_data['row_data'].columns:
            write_excel_col(
                workbook=workbook,
                sheet=sheet,
                data=table_data['row_data'].select(item),
                row_start=locations.get_row('end_row_header') + 1,
                col_start=locations.get_col('start_col_header_lhs') + i,
                base_style=styles.cell_rownames,
                data_styles=styles.data_styles,
            )
            i += 1

        if styles.merge_rownames:
            merge_rownames(
                workbook=workbook,
                sheet=sheet,
                table_data=table_data,
                locations=locations,
                styles=styles,
            )

    # Write the actual data itself
    if table_data['col_data'] is None:
        raise ValueError('Missing data')
    i = 0
    for item in table_data['col_data'].columns:
        write_excel_col(
            workbook=workbook,
            sheet=sheet,
            data=table_data['col_data'].select(item),
            row_start=locations.get_row('end_row_header') + 1,
            col_start=locations.get_col('start_col_header_rhs') + i,
            base_style=styles.cell_data,
            data_styles=styles.data_styles,
        )
        i += 1

    # Apply custom styles
    if styles.cell_styles is not None:
        for sty in styles.cell_styles:
            if not set(sty.cols).issubset(set(table_data['col_data'].columns)):
                raise ValueError(
                    f"Trying to style an element that was not found in the data: {[(c) for c in sty.cols if c not in table_data['col_data'].columns]}."
                )
            if any([r > table_data['col_data'].shape[0] for r in sty.rows]):
                raise ValueError(
                    'Trying to style a row outside of the range of the data.'
                )
            for col in sty.cols:
                for row in sty.rows:
                    set_region_style(
                        sheet=workbook[sheet],
                        style=sty.style,
                        start_row=locations.get_row('start_row_data') + row - 1,
                        start_col=locations.get_col('start_col_header_rhs')
                        + table_data['col_data'].columns.index(col),
                        end_row=locations.get_row('start_row_data') + row - 1,
                        end_col=locations.get_col('start_col_header_rhs')
                        + table_data['col_data'].columns.index(col),
                    )


def write_footnote(
    tbl: TableSpam,
    workbook: opy.Workbook,
    sheet: str,
    locations: Locations,
    styles: XlsxStyles,
) -> None:
    """Adds the footnote to the table.

    Args:
        tbl (TableSpam): TableSpam table created with TableSpam
        workbook (opy.Workbook): openpyxl workbook
        sheet (str, optional): name of the sheet to which the table should be added. Defaults to 'Table'.
        locations (dict[str, dict[str, int  |  None]]): a dict describing the locations (indexes) of different elements found in the table.
        styles (XlsxStyles): Styles that should be applied to the table.
    """
    if tbl.footnote is None:
        return

    loc = get_column_interval(
        start=locations.get_col('start_col_footnote'),
        end=locations.get_col('start_col_footnote'),
    )[0] + str(locations.get_row('start_row_footnote'))
    workbook[sheet][loc] = tbl.footnote
    workbook[sheet].merge_cells(
        start_row=locations.get_row('start_row_footnote'),
        start_column=locations.get_col('start_col_footnote'),
        end_row=locations.get_row('start_row_footnote'),
        end_column=locations.get_col('end_col_footnote'),
    )
    set_region_style(
        sheet=workbook[sheet],
        style=styles.cell_footnote,
        start_row=locations.get_row('start_row_footnote'),
        start_col=locations.get_col('start_col_footnote'),
        end_row=locations.get_row('start_row_footnote'),
        end_col=locations.get_col('end_col_footnote'),
    )


def create_outlines(
    tbl: TableSpam,
    workbook: opy.Workbook,
    sheet: str,
    locations: Locations,
    styles: XlsxStyles,
) -> None:
    """Adds vertical and horizontal lines in the table.

    Args:
        tbl (TableSpam): TableSpam table created with TableSpam
        workbook (opy.Workbook): openpyxl workbook
        sheet (str, optional): name of the sheet to which the table should be added. Defaults to 'Table'.
        locations (dict[str, dict[str, int  |  None]]): a dict describing the locations (indexes) of different elements found in the table.
        styles (XlsxStyles): Styles that should be applied to the table.
    """
    if tbl.header['lhs'] is not None:
        left_most = locations.get_col('start_col_header_lhs')
    else:
        left_most = locations.get_col('start_col_header_rhs')

    # top line
    set_region_style(
        sheet=workbook[sheet],
        style=styles.hline,
        start_row=locations.get_row('start_row_header'),
        start_col=left_most,
        end_row=locations.get_row('start_row_header'),
        end_col=locations.get_col('end_col_header_rhs'),
    )

    # bottom line
    set_region_style(
        sheet=workbook[sheet],
        style=styles.hline,
        start_row=locations.get_row('end_row_data') + 1,
        start_col=left_most,
        end_row=locations.get_row('end_row_data') + 1,
        end_col=locations.get_col('end_col_header_rhs'),
    )

    # left line
    set_region_style(
        sheet=workbook[sheet],
        style=styles.vline,
        start_row=locations.get_row('start_row_header'),
        start_col=left_most,
        end_row=locations.get_row('end_row_data'),
        end_col=left_most,
    )

    # right line
    set_region_style(
        sheet=workbook[sheet],
        style=styles.vline,
        start_row=locations.get_row('start_row_header'),
        start_col=locations.get_col('end_col_header_rhs') + 1,
        end_row=locations.get_row('end_row_data'),
        end_col=locations.get_col('end_col_header_rhs') + 1,
    )

    # row name separator
    set_region_style(
        sheet=workbook[sheet],
        style=styles.vline,
        start_row=locations.get_row('start_row_header'),
        start_col=locations.get_col('start_col_header_rhs'),
        end_row=locations.get_row('end_row_data'),
        end_col=locations.get_col('start_col_header_rhs'),
    )

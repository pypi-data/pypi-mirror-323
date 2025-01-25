import sys
from sys import flags
from typing import Tuple, Optional, Union, Callable
import io
import json
import mimetypes
from urllib.parse import unquote_plus
import os
import re
from collections import namedtuple
from datetime import datetime, timedelta, timezone
from pathlib import Path

import bs4
import pytz
import logging
import binascii
import rich.prompt
from prompt_toolkit.styles import Style

from result import Ok, Err, Result, is_ok, is_err
import attrs
from attrs import define, asdict
import canvasapi  # type: ignore
import requests
from openpyxl.styles.builtins import total
import lxml
from bs4 import BeautifulSoup, NavigableString
# from functools import lru_cache
try:  # type: ignore
    from pymemcache import serde  # type: ignore
    from pymemcache.client import base  # type: ignore

    MEMCACHED: Union[base.Client, bool] = True
except ImportError:
    print("No memcaching supported")
    MEMCACHED = False
#  from rich.logging import RichHandler
from rich.progress import track
from rich.progress import Progress
from rich.console import Console
from canvasapi.course import Course  # type: ignore
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment  # type: ignore
from canvasapi.util import combine_kwargs  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore
from openpyxl.workbook import Workbook  # type: ignore
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from socket import gaierror, timeout
from .canvasrobot_model import (AC_YEAR, NEXT_YEAR,  # type: ignore
                                EDUCATIONS, COMMUNITIES, LocalDAL, CanvasConfig,
                                EXAMINATION_FOLDER, COMMUNITY_EDU_IDS)  # type: ignore
from .entities import User, QuestionDTO, CourseMetadata, Grade, ExaminationDTO, \
    Stats  # type: ignore


class CustomConsole(Console):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

    def print(self, *args, **kwargs):
        if 'prefix' in kwargs and isinstance(args[0], str):
            l_args = [f"{kwargs['prefix']}{arg}" for arg in args]
            args = tuple(l_args)
            del kwargs['prefix']
        super().print(*args, **kwargs)


logging.getLogger("canvasapi").setLevel(logging.WARNING)
# we don't need the info messages
# from this library

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)


# noinspection PyProtectedMember
def file_update(file, **kwargs):
    """
        Updates a file in a course
        :calls: `PUT /api/v1/files/:id \
        <https://canvas.instructure.com/doc/api/files.html#method.files.update>`_
        :rtype: :class:`canvasapi.file.File`
        """
    response = file._requester.request(
        "PUT", "files/{}".format(file.id), _kwargs=combine_kwargs(**kwargs)
    )

    if "name" in response.json():
        response.set_attributes(response.json())

    return canvasapi.file.File(file._requester, response.json())


# define Python user-defined exceptions
class Error(Exception):
    """Base class for other exceptions"""
    pass


class DibsaRetrieveError(Error):
    """Raised when connection to Dibsa or retrieval failed"""
    pass


# noinspection PyClassHasNoInit
@define
class LabelType:
    label: str
    field_type: str


# noinspection PyClassHasNoInit
@define
class Profile:
    login_id: str
    id: int
    name: str
    short_name: str
    sortable_name: str
    avatar_url: str
    title: str
    bio: Optional[str]  # can be N
    primary_email: str
    integration_id: str
    time_zone: str
    locale: str
    # effective_locale str)
    # calendar str)
    # lti_user_id str)


@define
class Course2Foldername:
    course_id: int
    folder_name: str


if MEMCACHED:
    try:
        MEMCACHED = base.Client(('localhost', 11211),
                                connect_timeout=1,
                                timeout=0.2,
                                serde=serde.pickle_serde)
        result = MEMCACHED.set('canvasbot', 'Running')
        running = MEMCACHED.get('canvasrobot')
    except (ConnectionRefusedError, gaierror, timeout):
        MEMCACHED = False


# noinspection PyCallingNonCallable,GrazieInspection

def course_get_folder(course, folder_name: str):
    """"
    :course     Canvas course object
    :foldername Nane of the folder
    :returns    Canvas folder object

     """
    folders = course.get_folders()
    for folder in list(folders):
        if folder.full_name.endswith(folder_name):
            return folder
    return None


def course_metadata_memcached(course_id: int, canvas, ignore_assignment_names=None) -> CourseMetadata:
    """
    for course get the metadata from memcached if available. return CourseMetadata instance
    :param course_id
    :param canvas                     the canvas api object
    :param ignore_assignment_names  list of assignment_names to ignore in/for  the db
    :returns Course metadata instance
    """
    ignore_assignment_names = ignore_assignment_names or []

    def get_md():
        """ for statistical purposes:
        using the course_id:
        set canvas course var
        modules to point ar the canvas modules of this course
        'metadata vars about:
        - modules
        - quizzes
        - pages
        - assignments
        - submissions
        :returns course metadata object
        """
        course = canvas.get_course(course_id)
        modules = course.get_modules()
        nr_modules = len(list(modules))
        nr_module_items = 0
        nr_ext_urls = 0
        for module in modules:
            module_items = module.get_module_items()
            item_count_ext_url = filter(lambda i: i.type == "ext_url", module_items)
            nr_ext_urls += len(list(item_count_ext_url))
            nr_module_items += module.items_count
        pages = course.get_pages()
        pages = filter(lambda p: p.title[0:3] != 'UVT', pages)
        nr_pages = len(list(pages))

        assignments = course.get_assignments()
        assignments_summary = "Assignments:\n" if list(assignments) \
            else "No assignments"
        examinations = []
        for assignment in assignments:
            if assignment.name not in ignore_assignment_names:
                assignments_summary += f"{assignment.name}"
                examination = ExaminationDTO(course_id,
                                             course.name,
                                             assignment.name)

                examinations.append(examination)
            else:
                assignments_summary += f"ignored: {assignment.name}"
                # if show_all else ""

            submissions_summary = ""
            submissions = assignment.get_submissions()
            # try:
            for idx, submission in enumerate(submissions, start=1):
                if assignment.name not in ignore_assignment_names:
                    if submission.submission_type == "online_upload":

                        originality_str = (f"{submission.has_originality_report}"
                                           if hasattr(submission,
                                                      'has_originality_report')
                                           else "no Originality Report!")
                        submissions_summary += (f"({idx}. "
                                                f"{submission.submission_type}) "
                                                f"graded "
                                                f"{submission.grade} at "
                                                f"{submission.graded_at}. "
                                                f"Checked for plagiarism: "
                                                f"{originality_str}\n"
                                                )
                    else:
                        submissions_summary += (f"({idx}. "
                                                f"{submission.submission_type}) "
                                                f"graded "
                                                f"{submission.grade} at "
                                                f"{submission.graded_at}\n")
            assignments_summary += f"\n{submissions_summary}"
        nr_assignments = len(list(assignments))

        quizzes = course.get_quizzes()
        nr_quizzes = len(list(quizzes))

        all_files = course.get_files()
        nr_files = len(list(all_files))

        # check for uploaded examination files
        examination_files = 0
        examination_folder = course_get_folder(course, EXAMINATION_FOLDER)
        examinations_summary = "" if examination_folder \
            else f"No folder {EXAMINATION_FOLDER}"
        if examination_folder:
            files = examination_folder.get_files()
            for file in files:
                # folder = course.get_folder(file.folder_id)
                # if f"/{EXAMINATION_FOLDER}" in folder.full_name:
                # examination_folder_found = True
                examination_files += 1
                examinations_summary += f"{file.display_name}\n"

            examinations_summary += (f"\nTotal: {examination_files} "
                                     f"examination files ") if examination_files \
                else "No examination files"
        # was this working earlier 2.2.0 ?
        # try:
        #    collaborations = course.get_collaborations()
        #    list_of_cols = [c for c in collaborations]
        #    nr_collaborations = len(list_of_collaborations)
        # except TypeError:
        #    nr_collaborations = None
        cmd = CourseMetadata(nr_modules=nr_modules,
                             nr_module_items=nr_module_items,
                             nr_pages=nr_pages,
                             nr_assignments=nr_assignments,
                             nr_quizzes=nr_quizzes,
                             nr_files=nr_files,
                             assignments_summary=assignments_summary,
                             examinations_summary=examinations_summary,
                             examination_records=examinations
                             )
        return cmd

    if MEMCACHED:
        # memc_key = f"{course_id}{hash(ignore_assignment_names)}"
        memc_key = f"{course_id}c_m_c"
        md = MEMCACHED.get(memc_key)
        if md:
            return md
        MEMCACHED.set(memc_key, md := get_md())
        return md
    return get_md()


# noinspection PyTypeChecker,PyCallingNonCallable
class CanvasRobot(object):
    """" uses caching since """
    db: Callable
    canvas_login: bool = False
    TOT_WEIGHT: int = 100
    _year: int = AC_YEAR  # the current academic year
    outliner_foldernames: list[Course2Foldername] = []

    def __init__(self,
                 reset_api_keys: bool = False,
                 msg_queue=None,
                 console=None,  # optional Rich console
                 gui_root=None,  # optional Tkinter root
                 is_debug: bool = False,
                 is_testing: bool = False,
                 db_folder: Path = None,
                 db_auto_update: bool = False,
                 db_force_update: bool = False,
                 fake_migrate_all: bool = False):
        self.cookies: list = []
        self.browser = None
        self.queue = msg_queue  # for async use in tkinter
        self.console = console or CustomConsole()  # for commandline use
        self.gui_root = gui_root  # tkinter root
        self.is_debug = is_debug

        self.db_folder = db_folder if db_folder else Path.cwd() / 'databases'
        self.db = LocalDAL(is_testing=is_testing,
                           fake_migrate_all=fake_migrate_all,
                           folder=self.db_folder)

        config = CanvasConfig(reset_api_keys=reset_api_keys, gui_root=gui_root)
        try:
            self.canvas = canvasapi.Canvas(config.url, config.api_key)
            self.canvas_url = config.url
        except (canvasapi.exceptions.Forbidden, ConnectionError) as e:
            self.console.log(f"login Canvas failed ({e}) Connection trouble or wrong API key?")
            self.canvas_login = False
        else:
            if not self.canvas:
                self.console.log("login Canvas failed")
                self.canvas_login = False
        try:
            self.admin = self.canvas.get_account(config.admin_id) if config.admin_id \
                else None
        except (canvasapi.exceptions.Forbidden, ConnectionError, Exception) as e:
            self.admin = None
            self.console.log(f"Warning {e} no admin account in config?")

        db = self.db

        if db_force_update:
            self.update_database_from_canvas()
        elif db_auto_update:
            # is an update needed?
            row = db(db.setting.id == 1).select().first()
            self.last_db_update = row.last_db_update if row else datetime(1, 1, 1,
                                                                          tzinfo=timezone.utc)
            now = datetime.now(timezone.utc)
            utc_timezone = pytz.timezone('UTC')
            if getattr(self.last_db_update, 'tzinfo') is None:
                self.last_db_update = utc_timezone.localize(self.last_db_update).astimezone(utc_timezone)
            try:
                delta = now - self.last_db_update
            except (TypeError, Exception) as _:
                pass
            else:
                if db_auto_update and delta.days > 3:
                    self.update_database_from_canvas()

        self.teacher_ids = self.lookup_teachers_db()

        self.internal_id = None
        self.errors: list[str] = []
        self.actions: list[str] = []
        logger.info("Canvasrobot instance created")

    @property
    def year(self):
        return self._year

    @year.setter
    def year(self, year):
        self._year = year

    @property
    def full_ac_year(self):
        return f"{self._year}-{self._year + 1}"

    def add_to_queue(self, msg, value):
        if self.queue:
            self.queue.put((msg, value))

    def print(self, txt):
        if self.console:
            self.console.out(txt)
        else:
            print(txt)

    def report_errors(self):
        report = self.errors or "no errors recorded"
        if self.console:
            self.console.out(report)
        else:
            print(report)

    def print_outliner_foldernames(self) -> str:
        output = ""
        for item in self.outliner_foldernames:
            msg = (f"\n{self.canvas_url}/{item.course_id}"
                   f"/files/folder/{item.folder_name}")
            self.print(msg)
            output += msg
        return output

    def add_message(self, msg, value):
        # if gui (tkinter) send msg to queue
        if self.queue:
            self.queue.put((msg, value))
        if self.console:
            if self.is_debug:
                self.console.print(f"{msg}:{value=}" if value else msg)

    # COURSES----------------------------------------
    def get_course(self, course_id: int):
        """""
        :param course_id
        :returns canvas course by its id"""
        return self.canvas.get_course(course_id)

    def get_courses(self, enrollment: dict):
        """"
        :param enrollment 'teacher'(default), 'student','ta', 'observer' 'designer'
        :returns canvas courses for current user in role"""
        enrollment = enrollment or {'type': 'TeacherEnrollment'}
        return self.canvas.get_courses(enrollment=enrollment)

    def get_courses_in_account(self,
                               by_teachers: Optional[list] = None,
                               this_year=True):
        """
        get all courses from canvas in account.

        :param by_teachers: list of teacher id's
        :param this_year: True=filter courses to include only the current year
        :returns list of courses
        """
        console = self.console
        if not self.admin:
            console.log("No (sub) admin account provided")
            return []

        console.print("Get all courses, filter on current courses",
                      prefix='[green]‣[/green] ')
        with Progress(console=console) as progress:
            task_count = progress.add_task("[green]Getting all courses...",
                                           total=None)
            filtered_courses = []
            courses = self.admin.get_courses(by_teachers=by_teachers)
            total = len(list(courses))
            progress.remove_task(task_count)

            task_filter = progress.add_task(f"[green]Filter {total} courses...",
                                            total=total)
            for course in courses:
                progress.update(task_filter,
                                advance=1)  # Update de voortgangsbalk
                # only show/insert/update course if current year
                if this_year and (str(course.sis_course_id)[:4] != str(self.year)
                                  or course.name.endswith('conclude')):
                    continue
                filtered_courses.append(course)
        console.print(f"Number of current courses: {len(filtered_courses)}",
                      prefix='[green]‣[/green] '
                      )
        return filtered_courses

    def course_metadata(self, course_id, ignore_assignment_names=None):
        """
        return dict with metadata of this course
        as a dict.
        ignore the submissions of assignments present in ignore_assignment_names
        the others are reported in assignments_summary
        :returns md: CourseMetadata"""
        ignore_assignment_names = ignore_assignment_names or []
        md_result = course_metadata_memcached(course_id, self.canvas,
                                              frozenset(ignore_assignment_names))
        return md_result

    def get_all_active_courses(self, from_db=True):
        """:returns list of all canvas courses for this subadmin account"""

        def cur_year_active(course):
            """ filter function"""
            return (str(course.sis_course_id)[:4] == str(self.year) and
                    not course.name.endswith('conclude'))

        if from_db:
            db = self.db
            courses = db(
                (db.course.ac_year == self.year)).select(db.course.ALL)
            # map(set_id_to_course_id, courses)
        else:
            courses = self.admin.get_courses()
            courses = filter(cur_year_active, courses)
            logger.info("Create filtered list")
            courses = list(courses)
        return courses

    def get_course_using_osiris_id(self, osiris_id):
        """
        :returns first TST course with sisid starting with
        osiris_id in selected  year"""
        for course in self.admin.get_courses():
            # only consider course if selected year
            if str(course.sis_course_id)[:4] != str(self.year):
                continue
            if course.course_code.startswith(osiris_id):
                return course

    def get_course_id_using_osiris_id_from_db(self, osiris_id: str):
        db = self.db
        rows = db((db.course.course_code == osiris_id) &
                  (db.course.ac_year == self.year)).select(db.course.course_id)
        if rows:
            return rows[0].course_id
        return 0

    # COURSE PARTS: Modules, Pages
    def show_modules(self, course_id):
        course = self.get_course(course_id)
        modules = course.get_modules()
        # print(dir(modules))
        for page in modules:
            print(dir(page))
            print("page:{}".format(page))
            module_items = page.get_module_items()
            print("page module_items:{}".format(module_items))
            for item in module_items:
                print(item.c)
            # print("There are {} modules in page".format(len(page)))
        return

    def course_get_pages(self, course_id):
        """
        :param course_id
        :returns all all pages"""
        course = self.canvas.get_course(course_id)
        pages = course.get_pages(include=['body'])
        pages = filter(lambda p: p.title[0:3] != 'UVT', pages)  # skip
        return pages

    # USER, ROLES and PROFILE ----------------------------------------
    def get_user(self, user_id: int):
        """
        get user using
        :param user_id:
        :returns user
        """
        return self.canvas.get_user(user_id)

    @staticmethod
    def create_profile(profile_dict):
        """
        based on :profile_dict
        :returns Canvas profile or TypeError
        """
        try:
            profile = Profile(**profile_dict)
        except TypeError as e:
            logger.error(f"attribute needs to be added to Profile class: {e}")
            raise
        return profile

    def get_user_profile_id(self, user_id: int):
        user = self.get_user(user_id)
        profile = self.create_profile(user.get_profile())
        return user, profile

    def get_user_profile_anr(self, anr):
        user = self.canvas.get_user(anr, 'sis_user_id')
        profile = self.create_profile(user.get_profile())
        return user, profile

    def is_teacher_canvas(self, user):
        """
        check if user is a teacher in one of the TST courses by checking
        the canvas courses.
        :param user
        """
        role_teacher = {'type': 'TeacherEnrollment'}
        for course in self.admin.get_courses():
            enrollments = course.get_enrollments()
            for enrollment in enrollments:
                if enrollment.role == role_teacher and enrollment.user == user:
                    return True
        return False

    # get from DB
    def is_teacher_db(self, user):
        """" check if user is teacher in one of the TST courses by checking
        the database
        :param user"""
        return user.id in self.teacher_ids

    def lookup_teachers_db(self):
        db = self.db
        teachers = db((db.course2user.role == 'T') &
                      (db.course2user.user == db.user.id)).select(db.user.user_id,
                                                                  distinct=True)
        teacher_ids = [teacher.user_id for teacher in teachers]
        return teacher_ids

    # COURSE-USER interactions

    def search_user(self, search_name: str, email: str = ""):
        """
        :param search_name:
        :param email to search on email
        try search name as a login, then the email if supplied otherwise use
        """
        try:
            user = self.canvas.get_user(search_name, 'sis_login_id')
        except (canvasapi.exceptions.Unauthorized,
                canvasapi.exceptions.ResourceDoesNotExist) as e:
            if not email:
                # out of options
                self.errors.append(f'{e} And unable to lookup {search_name} '
                                   f'using email, parameter not provided')
                return False
            try:
                user = self.canvas.get_user(email, 'email')
            except canvasapi.exceptions.ResourceDoesNotExist:
                self.errors.append(f'{search_name} not found, {email} not found.')
                return False  # give up
            except canvasapi.exceptions.Unauthorized:
                self.errors.append(f"Using {search_name} "
                                   f"not (allowed to be) found by email in Canvas")
                return False  # give up
        return user

    # interact with EXTERNAL systems
    @staticmethod
    def get_students_dibsa(c_name, local=True):
        """get list of students from (local) DIBSA CRM"""
        # idea:  use LDAP instead
        url = (f"{'http://127.0.0.1:8000' if local else 'https://webapp.fkt.uvt.nl/'}"
               f"/dibsa/service/call/json/students/{c_name}")
        try:
            r = requests.get(url)
        except requests.exceptions.ConnectionError:
            raise DibsaRetrieveError(f"Unable to connect to source "
                                     f"Dibsa @ {url} maybe first start web2py locally?")
        if r.status_code != 200:
            raise DibsaRetrieveError(r.text)
        try:
            json_result = r.json()
        except json.decoder.JSONDecodeError:
            raise DibsaRetrieveError(f"unable to decode student data "
                                     f"of {c_name} from Dibsa @ {url}")
        except Exception as e:
            raise DibsaRetrieveError(f"unable to decode student data "
                                     f"of {c_name} from Dibsa @ {url} due to {e}")

        return json_result

    def get_students_for_community(self, c_id, local=False):
        """given
        :param local: get student dat from local Dibsa if True
        :param c_id community id
        lookup the list of educations (at least one item)
        retrieve the students from local dibsa/ldap
        for each student lookup canvas user and build list of canvas
        user (student)
        :returns dict  with as key the edu_id list of Canvas students"""

        community_edu_ids = COMMUNITY_EDU_IDS
        edu_ids = community_edu_ids[c_id]  # list of 1 or more items
        students_dibsa = {}
        for edu_id in edu_ids:
            try:
                students_dibsa[edu_id] = self.get_students_dibsa(edu_id.upper(),
                                                                 local=local)
            except DibsaRetrieveError as e:
                msg = f"Unable to retrieve students for {edu_id}"
                logger.error(e)
                raise DibsaRetrieveError(msg)
            if isinstance(students_dibsa, str) or hasattr(students_dibsa, "msg"):
                msg = students_dibsa["msg"] if hasattr(students_dibsa, "msg") \
                    else f"Unable to retrieve students for {edu_id}"
                logger.error(msg)
                raise DibsaRetrieveError(msg)

        students_canvas = {}
        for edu_id, students in students_dibsa.items():
            tmp_students = []
            first_enroll = []
            for student in students:
                try:
                    username = student['username']
                except TypeError as e:
                    logger.error(e)
                    self.errors.append(e)
                    continue

                user = self.search_user(username)
                if not user:
                    self.errors.append(f"{username} not found in Canvas,")
                    # todo: cn be reared by adding student to a or the course
                    first_enroll.append(user)
                    continue
                tmp_students.append(user)
            students_canvas[edu_id] = tmp_students, first_enroll
        # return a dict with the ed_id as key(s) and a list of studenten as values
        return students_canvas

    def get_community(self, c_id):
        # c_id can be an education (each education has a community)
        # or acskills (all educations)
        # Decide: add pm-ma to BA?
        #         pm-ulo to ULO?
        #         pm-macs to MACS
        try:
            course = self.get_course(COMMUNITIES[c_id])
        except IndexError:
            return None, [f"wrong c_id:{c_id}"]

        return course

    def enroll_in_course(self,
                         search: str = "",
                         course_id: int = 0,
                         section_id: int = 0,
                         username: str = "",
                         enrollment=None) -> Result[User, str]:
        """
        search course on osiris_id with enroll using username (more robust,
        using osiris_id if needed)

        :param search: if given get course using osiris_id
        :param course_id: course_id
        :param section_id: enroll in section if provided
        :param username:
        :param enrollment:
        enrollment={'type': 'StudentEnrollment'}
        :returns: result of course enrollment as a str...
        """
        if enrollment in (None, {}):
            enrollment = dict(type="StudentEnrollment")
        if search:
            try:
                course = self.get_course_using_osiris_id(search)
            except canvasapi.exceptions.ResourceDoesNotExist:
                return Err(f"Course {search} not found in Canvas")
        else:
            course = self.get_course(course_id)

        section = None
        if section_id:
            section = course.get_section(section_id)

        enrollment = enrollment or {'type': 'StudentEnrollment'}

        user = self.search_user(username)
        if not user:
            # search_user / get_user fails to find a newly imported account
            # work around it by using enroll_user
            try:
                if section:
                    section.enroll_user(f"sis_login_id:{username}")
                else:
                    course.enroll_user(f"sis_login_id:{username}", enrollment=enrollment)
            except Exception as e:
                msg = f"Failed with {e} while using special syntax for sys_login_id"
                self.errors.append(msg)
                return Err(msg)
            else:
                msg = f"Enrolled {username} in {course} as {enrollment}"
                self.actions.append(msg)
                return Ok(user)

        # user is found normally
        try:
            if section:
                section.enroll_user(user, enrollment=enrollment)
            else:
                course.enroll_user(user, enrollment=enrollment)
        except Exception as e:
            msg = f"Enroll of {user} in {course} as {enrollment} failed:{e}"
            self.errors.append(msg)
            return Err(msg)
        msg = f"Enrolled {user} in {course} as {enrollment}"
        self.actions.append(msg)
        return Ok(user)

    def enroll_user_in_course(self, user, course, role):
        """
        enroll identified user in course with role
        :param user:
        :param course:
        :param role:
        :return:
        """
        try:
            course.enroll_user(user, role)
        except (canvasapi.exceptions.BadRequest,
                canvasapi.exceptions.Conflict) as e:
            self.errors.append(f'User {user.name} not added to {course.name}: {e}')
        else:
            self.actions.append(f"{user.name} added to {course.name} as {role}")
        return

    def enroll_user_in_course_section(self, user, course, section_id, role):
        section = course.get_section(section_id)
        """
        enroll (valid) user in section of course with role
        :param user:
        :param course:
        :section_id:
        """
        try:
            section.enroll_user(user)
        except (canvasapi.exceptions.BadRequest,
                canvasapi.exceptions.Conflict) as e:
            self.errors.append(f'User {user.name} not added to {course.name} in section{section.name}: {e}')
        else:
            self.actions.append(f"{user.name} added to {course.name} in section {section.name} as {role}")
        return

    def enroll_students_in_communities(self, local=False):
        """for each community
        retrieve the students from local or central dibsa(ldap)
        and enroll them"""

        community_edu_ids = COMMUNITY_EDU_IDS
        for c_id in community_edu_ids:
            # for each community course
            # if c_id != "bauk":
            #    continue
            combi_dict = self.get_students_for_community(c_id,
                                                         local=local)
            course_id, lookup_sections = COMMUNITIES[c_id]
            course = self.get_course(course_id)
            role_student = {'type': 'StudentEnrollment'}
            # first handle the sis_login_ids
            for edu, combi in combi_dict.items():
                _, usernames = combi
                for username in usernames:  # students

                    if lookup_sections and edu in lookup_sections:
                        section = lookup_sections[edu]
                        section.enroll_user(f"sis_login_id:{username}")
                    else:
                        course.enroll_user(f"sis_login_id:{username}",
                                           dict(type="StudentEnrollment"))

                        # self.enroll_user_in_course(student, course, role=role_student)

            for edu, combi in students_dict.items():
                students, _ = combi
                for student in students:
                    if lookup_sections and edu in lookup_sections:
                        section = lookup_sections[edu]
                        self.enroll_user_in_course_section(student, course, section, role=role_student)
                    else:
                        self.enroll_user_in_course(student, course, role=role_student)

    def add_observer_to_education(self, user, report_only=False):
        """ add user as an observer to all courses of an education"""
        # todo: select courses using membership of education using osiris ids
        # not working just showing
        if not report_only:
            print("Add is not implemented!")
        print('Showing ALL observers for ALL courses')
        print(user)
        # idea: filter course of an education using db
        for course in self.admin.get_courses():
            # only insert/update course if current year
            if str(course.sis_course_id)[:4] != str(self.year):
                continue
            observers = course.get_users(enrollment={'type': 'ObserverEnrollment'})
            for observer in observers:
                if not hasattr(observer, 'login_id'):
                    continue
                print(observer.name)

    def remove_observer_from_all_courses(self, username):
        """ remove user as an observer from all TST courses"""
        removed = []
        try:
            user = self.canvas.get_user(username, 'sis_login_id')
        except canvasapi.exceptions.ResourceDoesNotExist:
            return f"User {username} not found in Canvas"

        for course in self.admin.get_courses():
            # only get a course if current year
            if str(course.sis_course_id)[:4] != str(self.year):
                continue
            enrollments = course.get_enrollments(type="ObserverEnrollment")
            # specifying user_id is not allowed
            for enrollment in enrollments:
                if enrollment.user_id != user.id:
                    continue
                print(enrollment)
                enrollment.deactivate(task='delete')
                removed.append(course.name)

        return removed

    def unenroll_in_course(self,
                           course_id: int = 0,
                           username: str = "",
                           enrollment=None) -> Result[User, str]:
        try:
            course = self.get_course(course_id)

            filter_enrollment = enrollment or {'type': 'StudentEnrollment'}

            user = self.search_user(username)

            enrollments = course.get_enrollments(enrollment=filter_enrollment)
            # doesn't filter !
            # specifying user_id is not allowed
            for enrollment in enrollments:
                if enrollment.user_id != user.id or enrollment.type != filter_enrollment['type']:
                    continue
                # print(enrollment)
                enrollment.deactivate(task='delete')
        except (canvasapi.exceptions.BadRequest,
                canvasapi.exceptions.Conflict) as e:
            msg = f'User {username} not removed to {course_id}: {e}'
            self.errors.append(msg)
            return Err(msg)
        else:
            self.actions.append(f"{username} removed from {course_id} as {filter_enrollment}")
            return Ok(user)

    def import_courses(self, filename):
        """from csv file updates table Course and Teacher
        :param filename: filename csv file
        """

        import csv
        from os.path import expanduser

        path = os.path.join(expanduser('~'), 'Downloads', filename + '.csv')
        db = self.db
        try:
            ofile = open(path, "rU")
        except IOError as e:
            logger.error("{} {} not found".format(e.message, path))
            raise
        else:
            reader = csv.reader(ofile, delimiter=';', quotechar='"')
            header = next(reader)
            logger.debug(header)
            # db(db.teacher.id > 0).delete()
            db(db.course2teacher.id > 0).delete()
            # refresh all course_teacher couplings
            for row in reader:
                logger.debug(row)
                assert len(row) == 7, "Error {0} fields!".format(len(row))
                teacher_names = row[2].replace(' (T)',
                                               '').replace(' (U)',
                                                           '').replace(' en ',
                                                                       ', ').split(', ')
                # phases = row[4].split(', ')
                course_id = db.course.update_or_insert(db.course.course_base == row[0],
                                                       course_base=row[0],
                                                       name=row[1],
                                                       teacher_names=row[2],
                                                       ects=int(row[3]),
                                                       phase=row[4],
                                                       department=row[5],
                                                       memo=row[6])
                if not course_id:  # must be there
                    course_id = db(db.course.course_base ==
                                   row[0]).select(db.course.id)[0].id
                else:
                    logger.info("course added")
                for t_name in teacher_names:
                    teacher_id = db.teacher.update_or_insert(db.teacher.name == t_name,
                                                             name=t_name)
                    if not teacher_id:
                        teacher_id = db(db.teacher.name ==
                                        t_name).select(db.teacher.id)[0].id
                    db.course2teacher.insert(course=course_id,
                                             teacher=teacher_id)

            db.commit()
            ofile.close()
        return

    # ----------------------------------------
    # Other interactions

    # # Old Blackboard interactions
    # def execute_command(self, command, params=None):
    #     """
    #     not in use ?
    #     :param command: command to execute
    #     :param params: parameters to tune command
    #     """
    #     # self.open()  # open canvas object
    #     try:
    #         url = self.base_url + (self.commands[command].format(params)
    #                                if params else self.commands[command])
    #     except Exception:
    #         raise NotImplementedError(f"error in command
    #         {command} or params {params}")
    #     else:
    #         try:
    #             self.browser.get(url)
    #         except Exception as e:
    #             raise ValueError("self.browser.get uses url {} e:{}".format(url, e))
    #         return self.browser
    #
    # # the commands ------------------------------------------------------------
    #
    # def ud_course_users(self):
    #     total_count = 0
    #     for bb_id in self.get_course_bb_ids():
    #         count, count_students = self.ud_students(bb_id=bb_id)
    #         total_count += count
    #
    #     return total_count
    #
    # def ud_students(self, bb_id=None):
    #     """ add enrolled users for given or currently selected course to db
    #     :param bb_id: if not given use self.internal_id
    #     """
    #
    #     db = self.db
    #     if bb_id:
    #         # set self,internal_id using lookup
    #         self.internal_id = bb_id
    #     else:
    #         assert self.internal_id
    #     # we need the primary key
    #     try:
    #         bbcourse_id = db(db.bbcourse.bb_id ==
    #                          self.internal_id).select(db.bbcourse.id).first().id
    #     except AttributeError:
    #         logger.info(f'course_id {self.internal_id} not found in our database!')
    #         return -1
    #
    #     logger.info(f'updating users for course_id {bbcourse_id}')
    #     # show overview enrolled users
    #     self.execute_command('enrolled_users', self.internal_id)
    #     # assumes course is selected
    #     count = 0
    #     count_students = 0
    #     # connect user to this course in db
    #     db.commit()
    #
    #     return count, count_students

    def get_courses_data(self):
        """
        for all courses: get a matrix and labels course name
        and other fields from db
        :return: features-matrix, labels """

        db = self.db  # cosmetic reasons
        qry = db.course.ac_year == self.year
        courses = db(qry).select(db.course.course_id,
                                 db.course.nr_modules,
                                 db.course.nr_module_items,
                                 db.course.nr_pages,
                                 db.course.nr_assignments,
                                 orderby=db.course.id)
        labels = [c.course_id for c in courses]
        features = [(c.nr_modules,
                     c.nr_module_items,
                     c.nr_pages,
                     c.nr_assignments) for c in courses]
        columns = ("nrModules", "nrModuleItems", ",nrPages", "nr_assignments")
        return features, labels, columns

    # def get_bbcourses(self, single_course: str = ""):
    #     """ for all courses: get coursename and other fields from db
    #     :param single_course: used for testing
    #     :return: rows/list of dicts """
    #
    #     db = self.db  # cosmetic reasons
    #     suffix = '-{}-'.format(self.year)
    #     qry = db.bbcourse.course_suffix.contains(suffix)
    #     if single_course:
    #         qry &= db.bbcourse.course_base == single_course
    #     rows = db(qry).select(db.bbcourse.bb_id,
    #                           db.bbcourse.name,
    #                           orderby=db.bbcourse.name)
    #     return rows

    def report_studentcounts(self):
        """ for all courses: get coursename and student count
        :return: list of dicts, with keys coursename, count"""
        # join bbcourses with users
        db = self.db  # cosmetic reasons
        count = db.bbuser.id.count()
        rows = db((db.course2user.bbcourse_id == db.bbcourse.id) &
                  (db.course2user.bbuser_id == db.bbuser.id) &
                  (db.course2user.role == 'S')).select(db.bbcourse.name,
                                                       count,
                                                       groupby=db.bbcourse.name,
                                                       orderby=db.bbcourse.name)
        for row in rows:
            row.count = row[count]
        return rows

    # def update_all_file_urls(self, max_courses: int = 0, single_course=None) -> str:
    #     """ for all courses harvest the attachments/files urls from Bb
    #     :param single_course: for testing
    #     :param max_courses:
    #     :returns list of found files
    #     """
    #     courses = self.get_courses(single_course=single_course)
    #     logger.info("{} courses to scan".format(len(courses)))
    #
    #     total_files: Optional[list] = []
    #     for count, course in enumerate(courses):
    #         if max_courses and count == max_courses:
    #             logger.info("Stopped after {} courses".format(max_courses))
    #             break
    #         logger.info("{}/{}:{}".format(count + 1, len(courses), course.name))
    #         self.get_course(course.id)  # check this!
    #         # areas = self.get_areas()  # top level areas (menu buttons)
    #         # logger.info("#{} areas#".format(len(areas)))
    #         # for area in areas:
    #         #     self.goto_area(area)  # select content area
    #         #     bb_files = self.get_files_from_area(level=0, area_name=area.name)
    #         #     # above function is recursive
    #         #     logger.info("{} files#".format(len(bb_files)))
    #         #     # test_file_name = bb_files[0].fname
    #         #     self.update_documents(bb_files)  # record data files in database
    #         #     total_files += bb_files
    #     return "{} courses scanned {} files found".format(
    #     len(courses),
    #     len(total_files) if total_files else 0)

    # the delegates ---
    # noinspection PyRedeclaration
    def make_enroll_file(self, courseid):
        self.db.make_enroll_file(courseid)

    def _check_bb(self, suffix=NEXT_YEAR):
        self.db.check_bb(suffix)

    # def show_user(self, user_id):
    #     self.execute_command('showuser', user_id)
    #     time.sleep(120)

    def update_documents_db(self, files):
        """ insert file object properties (fname, url) in table document
        :param files: list of file objects ( containing fname, url)
        """
        db = self.db
        for c_file in files:
            # lookup bbcourse id
            course_id = db(db.course.course_id ==
                           self.internal_id).select(db.course.id).first().id
            try:
                if c_file.level > 1:
                    db.document.update_or_insert(db.document.url == c_file.url,
                                                 url=c_file.url,
                                                 name=c_file.name,
                                                 course_id=course_id,
                                                 area=c_file.area_name,
                                                 check_status=0)
                else:
                    db.document.update_or_insert(db.document.url == c_file.url,
                                                 url=c_file.url,
                                                 name=c_file.name,
                                                 course_id=course_id,
                                                 area=c_file.area_name)
            except Exception as e:
                logger.error(
                    "'{0}' error in update_documents() while inserting "
                    "'{1}' in table Document".format(e, c_file.name))
            else:
                db.commit()

    def get_list_of_documents_db(self, course_id=None):
        """ get documents/attachments from db als a list
        """
        db = self.db
        if course_id:
            qry = ((db.course.ac_year == self.year) &
                   (db.document.course == db.course.id) &
                   (db.course.course_id == course_id))
        else:
            qry = ((db.course.ac_year == self.year) &
                   (db.document.course == db.course.id))

        try:
            items = db(qry).select(db.document.ALL, db.course.course_id, db.course.name)
        except Exception as e:
            logger.error("*{0} error in get_list_of_documents_db() "
                         "while selecting documents*".format(e))
            return []
        else:
            return items

    # noinspection PyUnresolvedReferences
    def download_documents_localfs(self):
        """ retrieve the real documents from LMS """
        db = self.db
        counters = namedtuple('Counters', ['total', 'ok', 'failed'])
        items = self.get_list_of_documents_db()
        counters.total = len(items)
        counters.ok = 0
        counters.failed = 0
        for item in items:
            status = self.download_file(fname=item.name, url=item.url)
            if status == 200:
                counters.ok += 1
            else:
                counters.failed += 1
            db(db.document.id == item.id).update(http_status=status)
        db.commit()
        return counters

    def transfer_file_to_server(self, url=None, fname=None):
        """download a file to uploads folder from external url using a temporary buffer
        :param fname: original name of file
        :param url: remote url to fetch file from
        """
        db = self.db
        filename = fname.strip()
        logger.debug("creating transfer buffer %s for %s" % (filename, url))
        # convert to fname/value pairs
        cookies = {}
        for cookie in self.cookies:
            cookies[str(cookie['fname'])] = cookie['value']
        #
        logger.info("cookies %s for %s" % (cookies, url))
        with io.BytesIO() as outfile:
            try:
                r = requests.get(url, cookies=cookies, stream=True)
            except requests.exceptions.RequestException as e:
                logger.error("couldn't get {} from file {} due to {}".format(filename,
                                                                             url,
                                                                             e))
                return None

            if r.status_code == 200:
                self.receive_file(db, filename, outfile, r, url)
            else:
                logger.info("html response {} url is now {}".format(r.status_code,
                                                                    r.url))
                # try again with redirected url (we assume redirection)
                r = requests.get(r.url, cookies=cookies, stream=True)
                if r.status_code == 200:
                    self.receive_file(db, filename, outfile, r, url)
                else:
                    logger.error('Failed {0}'.format(r))

        return r.status_code

    # noinspection PyUnresolvedReferences
    def transfer_files_to_server(self):
        """transfer all documents in the bbdocument table to the server"""
        db = self.db
        counters = namedtuple('Counters', ['total', 'ok', 'failed'])
        items = self.get_list_of_documents()
        counters.total = len(items)
        counters.ok = 0
        counters.failed = 0
        for item in items:
            status = self.transfer_file_to_server(fname=item.name, url=item.url)
            if status == 200:
                counters.ok += 1
            else:
                counters.failed += 1
            db(db.document.id == item.id).update(http_status=status)
        db.commit()
        return counters

    def is_leaf(self, tag: bs4.Tag):

        # one child: a string
        if (len(list(tag.children)) == 1 and  # the NavigableString is a child
                len(tag.contents) == 1 and
                isinstance(tag.contents[0], NavigableString)):
            return True
        # no siblings
        if (not tag.find_next_sibling() and
                not tag.find_previous_sibling() and
                len(tag.contents) == 0):
            return result
        # no children
        if (len(list(tag.children) == 0 and
                len(tag.contents) == 0)):
            return result

    # transformation functions
    def search_replace_in_page(self,
                               page: canvasapi.page,
                               search_term: str = "",
                               replace_term: str = "",
                               search_only: bool = True,
                               ignore_case: bool = True,
                               dryrun: bool = True) -> Tuple[int, str]:
        """
        :param ignore_case:
        :param page: Canvas page object to be searched or updated
        :param search_term: text to search
        :param replace_term: text to replace with
        :param search_only: if False Replace
        :param dryrun: no real replacement
        :returns:
        count (of occurrences), marked_body with found occurrences marked in red
        if not dryrun and not confirming replace count = 0

        """
        def check_leafs(tag: bs4.element.Tag):
            try:
                for ndx, child in enumerate(tag.children):
                    if not isinstance(child, bs4.NavigableString):
                        continue
                    # is NS...
                    if search_term_l in str(child).lower():
                        # child.string.replace_with(f"->{child.string}<-")
                        new_span = soup.new_tag("span", style='color:red;')
                        new_span.string = child.string
                        tag.contents[ndx] = new_span


            except (AttributeError, Exception) as e:
                Exception(f"mark_as_found failed {e}")

        flags = re.IGNORECASE if ignore_case else 0
        # if found inside a <a> link or inside an iframe: mark on the outside
        soup = BeautifulSoup(page.body, 'lxml')

        search_term_l = search_term.lower()
        # items = soup.find_all()
        items = list(soup.descendants)

        count = 0
        for item in items:
            # handle tags with links (href/src): wrap with a span
            if ((item.name == 'a' and search_term_l in str(item['href']).lower()) or
               (item.name == 'iframe' and search_term_l in str(item['src']).lower())):

                span = soup.new_tag("span", attrs={'style': "color:red;"})
                span.string = f"'{search_term}' in {item.name} ->"
                item.wrap(span)
                count += 1

            # handle plain text leafs
            if isinstance(item, bs4.NavigableString) and search_term_l in str(item).lower():
                # self.mark_as_found(item)
                # item.string = f""
                item_string, count = re.subn(search_term, '->' + search_term + '<-',
                                             item.string,
                                             flags=flags)
                item.string.replace_with(item_string)
                # parent = item.parent doesn't work for NS
                # pass
                # span = soup.new_tag("span", attrs={'style': "color:red;"})
                # parent.content[0].wrap(span)

        marked_body = str(soup)  # html
        # marked_body, count = re.subn(search_term, ' <span style="color: red;">' + search_term + '</span>',
        #                             page.body,
        #                             flags=flags)
        if not search_only and not dryrun:
            # let's only do this if you're really sure
            new_body, count = re.subn(search_term, replace_term, page.body, flags=flags)
            # update page in canvas
            if rich.prompt.Confirm.ask(f"Are you sure you want to replace '{search_term}' with "
                                       f"'{replace_term}' in "
                                       f"page<a href='{self.canvas_url}/courses/{page.course_id}/pages/{page.url}'"
                                       f"becoming '{new_body}' ?"):
                page.edit(wiki_page=dict(body=new_body))
            else:
                count = 0
        return count, marked_body


    def course_search_replace_pages(self,
                                    course_id: int,
                                    search_term: str = "",
                                    replace_term: str = "",
                                    search_only: bool = True,
                                    ignore_case: bool = True,
                                    dryrun: bool = True) -> Tuple[int, list[Tuple], str]:
        """
        In a course replace text (or html) in all pages
        :param ignore_case:
        :param course_id: canvas course_id
        :param search_term: text te find
        :param replace_term: text to replace with
        :param search_only: False is Replace
        :param dryrun: True: don't update just return marked body
        :returns tuple of count, list of page tuples, string of html bodies
        """
        total_count = 0
        marked_bodies = ""
        found_pages = []
        course = self.canvas.get_course(course_id)
        # check if the course is selected only because of an studentEnrollment
        if course.enrollments and len(course.enrollments) > 0:
            for enrollment in course.enrollments:
                if enrollment['role'] == 'StudentEnrollment':
                    return 0, [], ""  # effectively skipping

        pages = course.get_pages(include=['body'])
        pages = filter(lambda p: p.title[0:3] != 'UVT', pages)  # skip
        for page in pages:
            if page.body and search_term.lower() in page.body.lower():
                count, new_body = self.search_replace_in_page(page,
                                                              search_term=search_term,
                                                              replace_term=replace_term,
                                                              search_only=search_only,
                                                              ignore_case=ignore_case,
                                                              dryrun=dryrun)
                if count:
                    found_pages.append((course_id, course.name, page.url, page.title))
                total_count += count
                marked_bodies += new_body

        return total_count, found_pages, marked_bodies

    def course_search_replace_pages_all_courses(self,
                                                search_term: str,
                                                replace_term: str,
                                                search_only: bool = True,
                                                ignore_case: bool = True,
                                                dryrun: bool = False):

        with Progress(console=self.console) as progress:
            task_count = progress.add_task("[green]Getting all courses...",
                                           total=None)
            courses = self.canvas.get_courses()
            total = len(list(courses))
            progress.remove_task(task_count)

            task_filter = progress.add_task(f"[green]Search {total} courses...",
                                            total=total)

            sum_count, sum_found_pages, sum_marked_bodies = 0, [], ""
            for course in courses:
                progress.update(task_filter,
                                advance=1)  # Update de voortgangsbalk
                try:
                    count, found_pages, marked_bodies = self.course_search_replace_pages(course_id=course.id,
                                                                                         search_term=search_term,
                                                                                         replace_term=replace_term,
                                                                                         search_only=search_only,
                                                                                         ignore_case=ignore_case,
                                                                                         dryrun=dryrun)
                except (NameError, Exception) as e:
                    logger.error("Error {}".format(e))
                    pass
                else:
                    sum_count += count
                    if found_pages:
                        sum_found_pages.extend(found_pages),
                    sum_marked_bodies += marked_bodies

        return sum_count, sum_found_pages, sum_marked_bodies

    def get_examinations_from_database(self,
                                       single_course: int = None,
                                       orderby=None,
                                       candidate: bool = False):
        """ get all assignments/examinations in the current year
        """
        db = self.db
        # include the NULL values
        if single_course:
            qry = ((db.course.course_id == single_course) &
                   (db.examination.course == db.course.course_id) &  # join
                   (db.examination.candidate ==
                    candidate)) \
                if candidate else ((db.course.course_id == single_course) &
                                   (db.examination.course == db.course.course_id))
        else:
            qry = ((db.course.ac_year == self.year) &
                   (db.examination.course == db.course.course_id) &  # join
                   (db.examination.candidate ==
                    candidate)) if candidate else ((db.course.ac_year == self.year) &
                                                   (db.examination.course ==
                                                    db.course.course_id))
        orderby = orderby or db.course.course_code
        records = db(qry).select(db.examination.id,
                                 db.examination.course,
                                 db.examination.course_name,
                                 db.examination.name,
                                 db.examination.ignore,
                                 db.course.course_code,
                                 db.course.course_id,
                                 orderby=orderby)
        return records

    def get_courses_from_database(self,
                                  skip_courses_without_students=False,
                                  qry=None,
                                  orderby=None,
                                  fields=None):
        db = self.db
        if skip_courses_without_students:
            cur_qry = (db.course.nr_students > 0) & (db.course.ac_year == self.year)
        else:
            cur_qry = (db.course.ac_year == self.year)

        if qry:
            cur_qry = cur_qry & qry

        fields = fields or db.course.ALL
        orderby = orderby or db.course.course_code
        try:
            records = db(cur_qry).select(fields,
                                         orderby=orderby)
        except binascii.Error:
            logger.exception("Wrong content in image field?")
        else:
            return records

    def get_course_from_database(self, course_id):
        db = self.db
        record = db(db.course.course_id ==
                    course_id).select(db.course.ALL).first()
        return record

    def delete_course_from_database(self, course_id):
        db = self.db
        course_deleted = db(db.course.course_id == course_id).delete()
        examination_deleted = db(db.examination.course == course_id).delete()
        db.commit()
        return course_deleted and examination_deleted

    def update_record_db(self, qry, field, value):
        """ using a select query
        :param qry to find a record, update
        :param field with the
        :param value supplied"""
        db = self.db
        ud_fields = {field: value}
        # row = db(db[table][search_field] == search_id).select()
        # row2 = db(db.course.course_id == search_id).select()
        # db(db[table][search_field] == search_id).update(**ud_fields)
        db(qry).update(**ud_fields)
        db.commit()

    def update_db_for(self, course, only_course: bool = False):   # , single_course: int = None):
        """
        updating the course in the db
        :param only_course: if True we won't record examinations and documents
        :param course: a course object
        :return:
        """

        def count_students(course):
            # students
            students = course.get_users(enrollment={'type': 'StudentEnrollment'})
            # remove student with name=='Test Student'
            nr_students = len(list(filter(lambda s: s.name != 'Test student',
                                          list(students))))
            return nr_students

        db = self.db
        logger.debug("course: {}".format(course.name))

        # students
        nr_students = count_students(course)

        # ud teachers
        teacher_logins, teacher_names, teachers_ids = self.update_db_teachers(course)

        format_str = "%Y-%m-%dT%H:%M:%SZ"
        creation_date = datetime.strptime(course.created_at, format_str)

        examinations = (
            self.get_examinations_from_database(single_course=course.id))
        # filter: we want to ignore some examinations
        ignore_examination_names = [row.name for row in examinations
                                    if (row.course == course.id and row.ignore)]

        md = self.course_metadata(course.id, frozenset(ignore_examination_names))

        c_id = self.update_db_course(course, creation_date, md, nr_students, teacher_logins, teacher_names)

        if c_id:  # a new course has been inserted in the db: signal this
            db(db.course.id == c_id).update(status=2)
        else:
            c_id = db(db.course.course_id == course.id).select().first().id

        for item in [] if only_course else md.examination_records:
            # candidate is True if course_name in examination_list else False
            _ = db.examination.update_or_insert((db.examination.course ==
                                                 item.course_id) &
                                                (db.examination.name == item.name),
                                                course=c_id,
                                                course_name=item.course_name,
                                                name=item.name
                                                )
        for user_id in teachers_ids:
            db.course2user.update_or_insert((db.course2user.course == c_id) &
                                            (db.course2user.user == user_id),
                                            course=c_id,
                                            user=user_id,
                                            role='T')

        files = [] if only_course else course.get_files()
        for file in files:
            _ = db.document.update_or_insert((db.document.course == c_id) &
                                             (db.document.url == file.url),
                                             course=c_id,
                                             folder_id=file.folder_id,
                                             url=file.url,
                                             filename=unquote_plus(file.filename),
                                             size=file.size,
                                             content_type=getattr(file, 'content-type'))

        db.commit()

        return c_id  # course id in db for new of existing course

    def update_db_course(self, course, creation_date, md, nr_students, teacher_logins, teacher_names):
        # update table course
        db = self.db
        # course_id = None
        try:
            course_id = db.course.update_or_insert(db.course.course_id == course.id,
                                                   course_id=course.id,
                                                   # status=course.workflow_state,
                                                   course_code=course.course_code,
                                                   sis_code=course.sis_course_id,
                                                   # year course_code and suffixes
                                                   name=course.name,
                                                   creation_date=creation_date,
                                                   ac_year=self.year,
                                                   nr_students=nr_students,
                                                   nr_modules=md.nr_modules,
                                                   nr_module_items=md.nr_module_items,
                                                   nr_pages=md.nr_pages,
                                                   nr_assignments=md.nr_assignments,
                                                   nr_quizzes=md.nr_quizzes,
                                                   assignments_summary=md.assignments_summary,
                                                   examinations_summary=md.examinations_summary,
                                                   teachers=teacher_logins,
                                                   teachers_names=teacher_names)
            db.commit()  # really needed here??
        except Exception as e:
            err = f"{e} error inserting {course.name}"
            self.add_message("<InsertError>", err)
            logger.exception(err)
            raise
        return course_id

    def update_db_teachers(self, course) -> Tuple[list[str], list[str], list[int]]:
        db = self.db
        teachers = course.get_users(enrollment={'type': 'TeacherEnrollment'})
        teachers_ids = []
        for teacher in teachers:
            if not hasattr(teacher, 'login_id'):
                continue
            # print(teacher.name)
            first_name, last_name, prefix = self.parse_sortable_name(teacher)
            inserted_id = db.user.update_or_insert(db.user.username ==
                                                   teacher.login_id,
                                                   user_id=teacher.id,
                                                   name=teacher.name,
                                                   first_name=first_name,
                                                   prefix=prefix,
                                                   last_name=last_name,
                                                   username=teacher.login_id,
                                                   email=teacher.email,
                                                   role='T')
            inserted_id = inserted_id or db(db.user.username ==
                                            teacher.login_id).select().first().id
            teachers_ids.append(inserted_id)
        teacher_logins, teacher_names = [], []
        try:
            # skips teachers with non-accepted invites
            # (they don't have a login attribute)
            teacher_logins = [teacher.login_id for teacher in teachers
                              if hasattr(teacher, 'login_id')]
            teacher_names = [teacher.name for teacher in teachers
                             if hasattr(teacher, 'login_id')]
        except AttributeError as e:
            # no login_id
            msg = (f"skipped teacher of {course.name} "
                   f"[{course.id}] due to {e}")
            logger.warning(msg)
        logger.debug("instructors: {0}".format(teacher_logins))  # instructors
        return teacher_logins, teacher_names, teachers_ids

    # noinspection PyUnusedLocal
    def update_database_from_canvas(self,
                                    single_course=None,
                                    single_course_osiris_id=None,
                                    max_number=None,
                                    stop_list=None):
        """
            Use the canvasapi to read the courses for the selected year
            (or a single course using canvas course_id or the osiris id) and
            - record internal_id course_id, fname and instructors in the
            table course
            - put teacher details in table user
            - collects info about assignments in assignment_summary
            - collects info about examinations files in examination_summary
            - reports new tentamination candidates
            :param single_course
            :param single_course_osiris_id
            :param max_number: stop earlier
            :param stop_list: courses to ignore
            :return number of added/updated rows
            """

        db = self.db

        msg = f'Open single course {single_course}' \
            if single_course \
            else f'open course(s) for year {self.year} - {self.year + 1}'
        self.add_message(msg, None)
        logger.info(msg)

        if single_course_osiris_id:
            courses = [self.get_course_using_osiris_id(single_course_osiris_id)]
            target = f" course {single_course}"
        elif single_course:
            courses = [self.get_course(single_course)]
            target = f" course {single_course}"
        else:
            courses = self.admin.get_courses()
            target = " all courses"
        if max_number:
            target = f" {max_number} courses"

        with Progress(console=self.console) as progress:
            task_count = progress.add_task("[green]Counting courses...", total=None)

            num_rows = 0
            num_courses = len(list(courses))
            max_number = max_number or num_courses
            progress.remove_task(task_count)

            task_process = progress.add_task("[green]Process courses...",
                                             total=num_courses)
            for idx, course in enumerate(courses):
                progress.update(task_process,
                                description=f"[green]Processing course {course.name}...",
                                advance=1)  # Update de voortgangsbalk
                self.add_message("<Progress>", (course.name, idx, num_courses))

                # only insert/update course if current year unless single_course
                if (str(course.sis_course_id)[:4] != str(self.year)
                    or course.name.endswith('conclude')) \
                        and not (single_course or single_course_osiris_id):
                    continue
                # skip specified courses
                if stop_list and course.name in stop_list:
                    continue
                # break prematurely?
                if idx > max_number:
                    break

                self.update_db_for(course)  # , single_course=single_course)
                num_rows += 1

            msg = f"[green]Updated db from Canvas for {target}. {num_rows} rows changed"
            self.console.print(msg)
            self.add_message("<Done>", msg)
        # record date of update
        db.setting.update_or_insert(db.setting.id == 1,
                                    last_db_update=datetime.now(timezone.utc))
        db.commit()
        return num_rows

    def is_user_valid(self, userinfo):
        """"
        :param userinfo (dict, instance, named tuple or Storage instance with
        attributes id, login_id)
        :returns True for valid user, detail"""
        if isinstance(userinfo, dict):
            # import collections
            # User = collections.namedtuple('User', 'id')
            user = User(id=userinfo['id'])
        else:
            user = userinfo

        user = self.canvas.get_user(user.id)
        profile = user.get_profile()
        if 'login_id' not in profile.keys():
            return True, 'no login_id, assume valid'
        if 'closed' in profile['login_id']:
            return False, 'login closed'
        if 'invalid' in profile['primary_email']:
            return False, 'mail invalid'
        return True, 'appears valid'

    @staticmethod
    def valid_c_id(c_id):
        valid_c_ids = [edu for edu in EDUCATIONS]
        valid_c_ids.append('ACSKILLS')
        if c_id.upper() not in valid_c_ids:
            return False, [f"Wrong c_id '{c_id.upper()}' should "
                           f"be one of {valid_c_ids} in lower case"]
        return True, ""

    @staticmethod
    def deactivate_enrollment(enrollment):
        enrollment.deactivate('delete')

    def deactivate_if_invalid(self, enrollment, report_only):
        valid, reason = self.is_user_valid(enrollment.user)
        if valid:
            return "no action"
        user_name = enrollment.user['fname']
        if report_only:
            self.actions.append(f"{user_name} would be deactivated ({reason})")
            return "would be deactivated"
        try:
            self.deactivate_enrollment(enrollment)
        except Exception as e:
            self.errors.append(f"Deactivate of {user_name} failed:{e}")
            return "deactivation failed"
        else:
            self.actions.append(f"{user_name} deactivated ({reason})")
        return "deactivated"

    def get_enrollments(self, c_id):
        course = self.get_course(COMMUNITIES[c_id])
        enrollments = course.get_enrollments()

        # errors = []

        def filter_student(enrollment):
            """filter the real students"""
            try:
                is_real_student = (enrollment.user['name'].lower() != 'test student' and
                                   enrollment.role == 'StudentEnrollment')
            except Exception as e:
                self.errors.append(f"{enrollment.user['name']} not retrieved {e}")
                return False
            return is_real_student

        # enrollments = [enrollment for enrollment in enrollments]
        enrollments = filter(filter_student, enrollments)
        return enrollments

    def cleanup_community(self, c_id, report_only):
        """
        given an id of a community (course)
        remove invalid enrollments
        """
        # result = self.check_c_cid(c_id)
        # if result != 'ok':
        #    return result

        course = self.get_course(COMMUNITIES[c_id])
        enrollments = course.get_enrollments()
        errors = []
        removed = []
        for enrollment in enrollments:
            if enrollment.user['fname'].lower() == 'test student':
                continue
            if enrollment.role == 'StudentEnrollment':
                try:
                    valid, reason = self.is_user_valid(enrollment.user)
                except Exception as e:
                    errors.append(f"{enrollment.user['fname']} not retrieved {e}")
                    continue
                if not valid:

                    if not report_only:
                        try:
                            enrollment.deactivate('delete')
                        except Exception as e:
                            errors.append(e)
                            continue
                    removed.append((enrollment.user['fname'], reason))

        num_removed = len(removed)
        msg = (f"{num_removed} students would have been removed"
               if report_only
               else f"{num_removed} students removed")

        return removed, errors, msg

    @staticmethod
    def parse_sortable_name(user):
        """examples:
        Klein, Wim
        Groot, Nico de
        Govaert, Samuel (Sam)
        Wieringen, Archibald (H.M.J.)
        """

        assert ", " in user.sortable_name, "sortable_name should contain comma"
        source = user.sortable_name
        pat = re.compile(
            r'(?P<last_name>[\w \-]+), (?P<first_name>\w+)'
            r'\s?((\((?P<first_name_par>\w+)\))|'
            r'(\((?P<init>[\w.]+.)\)))?(\s*(?P<prefix>\w+))?')
        d = re.match(pat, source)
        if not d:
            print("!!!Failed parsing!!!")

        return d['first_name_par'] or d['first_name'], d['prefix'] or '', d['last_name']

    def download_file(self, url=None, fname=None):
        """download a remote file to static folder from url
        :param fname: local file to write to
        :param url: remote url to fetch file from
        """
        filename = 'static/' + fname.strip()
        logger.info("creating file %s for %s" % (filename, url))
        # convert to fname/value pairs
        cookies = {}
        for cookie in self.cookies:
            cookies[str(cookie['fname'])] = cookie['value']
        with open(filename, 'wb') as outfile:
            r = requests.get(url, cookies=cookies, stream=True)
            if r.status_code == 200:
                for block in r.iter_content(1024):
                    if not block:
                        break
                    outfile.write(block)
            else:
                logger.info("html error %s" % r.status_code)
        return r.status_code

    @staticmethod
    def receive_file(db, filename, outfile, r, url):
        """ receive downloading file in database"""
        logger.debug('Start receiving file in buffer')
        for block in r.iter_content(1024):
            if not block:
                logger.debug('.')
                break
            outfile.write(block)
        outfile.seek(0)  # rewind
        content_type = r.headers['content-type']
        full_filename = 'undef'
        try:
            full_filename = filename + mimetypes.guess_extension(content_type) or ''
            thisfile = db.bbdocument.bbfile.store(outfile, full_filename)
        except TypeError as e:
            msg = "unable to store {} with content-type {} {}".format(filename,
                                                                      content_type,
                                                                      e)
            logger.error(msg)
        except IOError as e:
            msg = "unable to store {} based on {} and {} {}".format(full_filename,
                                                                    filename,
                                                                    content_type,
                                                                    e)
            logger.error(msg)
        else:
            logger.info("{} should be in uploads folder...".format(full_filename))
            db(db.bbdocument.url == url).update(bbfile=thisfile)
            db.commit()

    def course_grades(self, c_id):
        course = self.get_course(c_id)
        enrollments = course.get_enrollments()
        grades = []
        for enrollment in enrollments:
            if enrollment.user['fname'].lower() == 'test student':
                continue
            if enrollment.role == 'StudentEnrollment':
                # print(enrollment)
                grades.append(Grade(stud_name=enrollment.user['fname'],
                                    stud_id=enrollment.user['login_id'],
                                    final_score=enrollment.grades['final_score'],
                                    final_grade=enrollment.grades['final_grade']))

        return course, grades

    def get_grades(self, c_id):
        """" From the course
        :param c_id
        get the grades and create an Excel file for Osiris import"""
        course, grades = self.course_grades(c_id)

        labels_and_types = [LabelType(label='stud_id', field_type='string'),
                            LabelType(label='stud_name', field_type='string'),
                            LabelType(label='score', field_type='percentage'),
                            LabelType(label='grade', field_type='grade')]
        styles = {
            # list below is not complete
            # see https://www.web2pyref.com/reference/field-type-database-field-types
            'datetime': NamedStyle(name='date', number_format='yyyy-mm-dd'),
            'id': NamedStyle(name='int', number_format='#,##0'),
            'integer': NamedStyle(name='int', number_format='#,##0'),
            'price': NamedStyle(name='money', font=Font(italic=True),
                                fill=PatternFill(fill_type='solid', fgColor='404040'),
                                number_format=u'[$\u20ac-1] #,##0.00 '),  # unicode €
            'double': NamedStyle(name='double', number_format='#0.000000'),
            'grade': NamedStyle(name='grade', number_format='#0.00'),
            'percentage': NamedStyle(name='score', number_format='#0.00'),
            'string': NamedStyle(name='std',
                                 alignment=Alignment(horizontal='left',
                                                     vertical='bottom',
                                                     text_rotation=0,
                                                     wrap_text=False,
                                                     shrink_to_fit=False,
                                                     indent=0)),
            'text': None
        }

        dest_filename = f'TST {course}.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToHeight = 0
        ws.page_setup.fitToWidth = 1
        # add 6 describing rows
        ws.append(["Faculteit", "", "TST"])
        ws.append(["Vakbenamimg", "", course.name])
        ws.append(["Vakcode", "", course.course_code])
        ws.append(["Datum tentamen", "", "[in te vullen]"])
        ws.append([])
        ws.append(["Student", "SIS User ID", "Final Score", "Final Grade"])
        # data
        for row in grades:
            ws.append(attrs.astuple(row))

        # set width
        dim_holder = DimensionHolder(worksheet=ws)
        widths = [25, 12, 12, 12]
        for index, col in enumerate(range(ws.min_column, ws.max_column + 1)):
            dim_holder[get_column_letter(col)] = ColumnDimension(ws,
                                                                 min=col,
                                                                 max=col,
                                                                 width=widths[index])
        ws.column_dimensions = dim_holder

        # apply the styles, must be cell by cel
        for row in ws.iter_rows(min_row=2):
            for index, cell in enumerate(row):
                try:
                    lookup = labels_and_types[index].field_type  # lookup
                    cell.style = styles[lookup]  # lookup
                except ValueError:
                    pass

        # mark first six row as header
        for cell in ws[6]:
            cell.style = 'Pandas'
        ws.freeze_panes = "A7"
        ws.print_title_rows = '1:6'
        wb.save(filename=dest_filename)

        return (f"Excel file {dest_filename} created "
                f"with {len(grades)} rows"), dest_filename

    def create_quiz(self, course_id: int, title: str, quiz_type: str = '') -> int:
        """
        :param course_id:
        :param title:
        :param quiz_type:
        :return: msg, quiz_id
        """
        course: Course = self.get_course(course_id)
        quiz = course.create_quiz(dict(title=title, quiz_type=quiz_type))
        logger.debug(f"{course.name} now contains {quiz}")
        return quiz.id

    def create_question(self,
                        course_id: int,
                        quiz_id: int,
                        question_dto: QuestionDTO) -> int:
        """
        :param course_id:
        :param quiz_id:
        :param question_dto:
        :return: msg, quiz_question_id
        """
        course = self.get_course(course_id)
        quiz = course.get_quiz(quiz_id)
        quiz_question = quiz.create_question(question=asdict(question_dto))
        logger.debug(f"{quiz} now contains {quiz_question}")
        return quiz_question.id
        # return f"{quiz} now contains {quiz_question}", quiz_question.id

    def create_quizzes_from_data(self,
                                 course_id: int,
                                 question_format="Question {}.",
                                 data=None,
                                 gui_root=None,
                                 gui_queue=None
                                 ):
        """
        :param course_id: Canvas course_id: the quizzes are added to this course
        :param question_format: used to create the question name.
        They will be numbered. Should contain '{}' as  placeholder
        starting with 1
        :param data: the quizdata
        :param gui_root: used in combination with GUI (tkinter)
        :param gui_queue: used in combination with GUI
        :return: stats:Stats
        """
        if '{}' not in question_format:
            msg = (f"parameter 'question_format(={question_format})' "
                   f"should contain {{}} als placeholder")
            ValueError(msg)

        stats = Stats()

        total_questions = 0
        for _, questions in data:
            total_questions += len(questions)

        for quiz_name, questions in track(data):  # for non-gui progress

            quiz_id = self.create_quiz(course_id=course_id,
                                       title=quiz_name,
                                       quiz_type="practice_quiz")
            stats.quiz_ids.append(quiz_id)
            for index, (question_text, answers) in enumerate(questions, start=1):
                # answers_asdict = [asdict(answer) for answer in answers]
                question_dto = QuestionDTO(question_name=question_format.format(index),
                                           question_text=question_text,
                                           answers=answers)
                question_id = self.create_question(course_id=course_id,
                                                   quiz_id=quiz_id,
                                                   question_dto=question_dto)
                stats.question_ids.append(question_id)
                if gui_root:
                    gui_queue.put(index / total_questions)
                    gui_root.event_generate('<<CreateQuizzes:Progress>>')

        if gui_root:
            gui_root.event_generate('<<CreateQuizzes:Done>>')

        return stats

    def get_course_tab_by_label(self, course_id: int, label: str):
        course = self.get_course(course_id)
        for tab in course.get_tabs():
            if tab.label == label:
                return tab

    def create_folder_in_course_files(self, course_id: int, foldername: str,
                                      locked=True, report_only=False):
        """
        :param report_only: defaults to False, if true the folder is not created
        :param locked: defaults to True, folder is unpublished and invisible
        for students
        :param foldername:
        :param course_id:
        :returns folder_id of True if report_only is True and a folder would be created
        in course with 'course_id' create a folder 'foldername'"""
        course = self.get_course(course_id)
        folder_created = 0

        def name_variant(source, target):
            # match Tentamens, Tentamens 2020, etc but not tentamens/inner
            # noinspection RegExpRedundantEscape
            result = re.match(fr'^course files\/{source}(?!.*\/.*$)', target)
            return result

        folders_named_exact = [folder for folder in course.get_folders()
                               if f"course files/{foldername}" == folder.full_name]
        folders_named_variants = [folder for folder in course.get_folders()
                                  if name_variant(foldername, folder.full_name)]
        for folder in folders_named_variants:
            # find/check folder(s) containing 'file_name' with possible postfixes
            # like f'{file_name} 2025'
            if folder.name != foldername:  # variant, postfix
                self.outliner_foldernames.append(Course2Foldername(course_id,
                                                                   folder.full_name))
            logger.info(f"No creation needed: Folder"
                        f" '(course) files/{foldername}' "
                        f"already in ({course_id=})")
            if not folder.locked and locked:
                logger.warning(f"Folder {folder.full_name} was not locked, locking it "
                               f"and its children now!")
                folder.update(locked=locked)
                _num_items = self.unpublish_folderitems_in_course(course_id,
                                                                  foldername=folder.full_name,
                                                                  files_too=True,
                                                                  check_only=False)
        if len(folders_named_exact) == 0:  # original folder simply not there
            if report_only:
                logger.info(f"Folder '(course) files/{foldername}' "
                            f"would be created now in ({course_id=})")
            else:
                folder_id = course.create_folder(foldername,
                                                 parent_folder_path='/',
                                                 locked=locked)
                logger.info(f"Folder '(course) "
                            f"files/{foldername}' ({folder_id}) is created "
                            f"as {locked=} now in ({course_id=})")
            folder_created += 1

        return folder_created

    def create_folder_in_all_courses(self, foldername,
                                     locked=True,
                                     report_only=False):

        folders_created = 0
        for course in track(self.get_all_active_courses(from_db=False),
                            description="All current courses...",
                            console=self.console):
            folders_created += self.create_folder_in_course_files(course.id, foldername,
                                                                  locked=locked,
                                                                  report_only=report_only)
        logger.info(f"{folders_created} folders"
                    f"{' would be' if report_only else ''} created")
        for course_id, foldername in self.outliner_foldernames:
            txt = f'Found {foldername} in https://{self.canvas_url}/{course_id}/files'
            self.print(txt)

    def unpublish_subfolder_in_all_courses(self,
                                           foldername: str,
                                           files_too: bool = False,
                                           check_only: bool = False):
        for course in track(self.get_all_active_courses(from_db=False),
                            description=(f"Checking all current"
                                         f" courses "
                                         f"for folder '{foldername}'..." if check_only
                            else f"Unpublish all published folder {foldername}..."),
                            console=self.console):
            if course.name.endswith("_conclude"):
                continue
            # logger.info(course.name)
            self.unpublish_folderitems_in_course(course.id,
                                                 foldername,
                                                 files_too,
                                                 check_only)

    def unpublish_folderitems_in_course(self, course_id: int,
                                        foldername: str,
                                        files_too: bool = False,
                                        check_only: bool = False):
        """
        :param check_only: only show publication status no change
        :param foldername:
        :param course_id:
        :param files_too: if true unpublish files in the folder too (recursive)
        :returns  list of course ids of courses with missing foldername
        """
        file_changes = 0
        folder_changes = 0

        def unpublish_items(file_or_folder):
            nonlocal folder_changes
            nonlocal file_changes
            nonlocal course_id
            for folder_ in file_or_folder.get_folders():
                if not folder_.locked:
                    folder_.update(locked=True)
                    folder_changes += 1
                    logger.info(f"Corrected: Folder '{folder_.full_name}' "
                                f"is now unpublished!")
                unpublish_items(folder_)
            for file in file_or_folder.get_files():
                if not file.locked:
                    if not check_only:
                        file_update(file, locked=True)
                        logger.info(f"Corrected: File "
                                    f"'{file.display_name}' in {foldername} "
                                    f"is now unpublished")
                        file_changes += 1
                    else:
                        logger.warning(f"File '{file.display_name}' "
                                       f"in {foldername} is published!")

        # files_folder = 'course files'
        course_ids_missing_folder = []
        course = self.get_course(course_id)
        folders = course.get_folders()
        for folder in folders:  # paginated
            if folder.full_name == foldername:
                files_tab = self.get_course_tab_by_label(course_id, "Files") or \
                            self.get_course_tab_by_label(course_id, "Bestanden")
                try:
                    if files_tab.visibility == "public":
                        logger.warning(f"Files folder of {course.name}"
                                       f" ({course.id}) is visible")
                        # files_tab.visibility = "admins"
                        # ! no change  without teachers approval!
                except AttributeError:
                    logger.warning(f"Files tab visibility of "
                                   f"{course.name} {course_id} missing")

                # folder_id = folder.id
                if not folder.locked:
                    if not check_only:
                        folder.update(locked=True)
                        folder_changes += 1
                        logger.info(
                            f"Folder '{foldername}' is now unpublished"
                            f" in course {course.name}")
                    else:
                        logger.warning(f"Folder '{foldername}' is published "
                                       f"in course {course.name}({course.id})!")
                else:
                    logger.debug(
                        f"Folder '{foldername}' was already "
                        f"unpublished/locked in course {course.name}")
                if files_too:
                    unpublish_items(folder)
                    # for f in folder.get_files():
                    #    if not f.locked:
                    #        logger.warning(f"{f.filename} is published!")
                break
        else:
            logger.info(f"Folder '{foldername}' not found in {course.name}")
            course_ids_missing_folder.append(course_id)

        for course_id in course_ids_missing_folder:
            logger.info(f"Folder '{foldername}' not found in {course_id}")
        return folder_changes, file_changes


if __name__ == '__main__':
    pass

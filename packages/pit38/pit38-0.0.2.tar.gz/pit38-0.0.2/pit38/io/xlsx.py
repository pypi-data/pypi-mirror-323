from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook

from pit38.io.utils import serialize_decimal
from pit38.models import ClosedPosition


def read_trades_from_xlsx(file: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    """
    Reads an XLSX file (entire sheet) into a list of dictionaries (raw data).
    The first row is assumed to be the header.
    """
    wb = openpyxl.load_workbook(file, data_only=True)
    sheet = wb[sheet_name] if sheet_name else wb.active

    rows = list(sheet.rows)
    if not rows:
        return []

    headers = [cell.value if cell.value else "" for cell in rows[0]]
    raw_data = []
    for row in rows[1:]:
        row_dict = {}
        for col_idx, cell in enumerate(row):
            if col_idx < len(headers):
                header = headers[col_idx].strip()
                row_dict[header] = cell.value
        raw_data.append(row_dict)
    return raw_data


def write_closed_positions_to_xlsx(
    profit_positions: list[ClosedPosition], loss_positions: list[ClosedPosition], file: Path
) -> None:
    """
    Writes the matched buy-sell trades into an XLSX file
    """
    wb = Workbook()
    wb.remove(wb.active)
    profit_ws = wb.create_sheet("Profit")
    loss_ws = wb.create_sheet("Loss")

    headers = [
        "ISIN",
        "Ticker",
        "Currency",
        "BuyDate",
        "Quantity",
        "BuyAmount",
        "BuyCommission",
        "BuyExchangeRate",
        "SellDate",
        "SellAmount",
        "SellCommission",
        "SellExchangeRate",
        "Profit",
        "IncomePLN",
        "CostsPLN",
    ]
    for ws in [profit_ws, loss_ws]:
        ws.append(headers)

    def create_row(pos: ClosedPosition) -> list:
        return [
            pos.isin,
            pos.ticker,
            pos.currency,
            pos.buy_date.strftime("%Y-%m-%d"),
            serialize_decimal(pos.quantity),
            serialize_decimal(pos.buy_amount),
            serialize_decimal(pos.buy_commission),
            serialize_decimal(pos.buy_exchange_rate),
            pos.sell_date.strftime("%Y-%m-%d"),
            serialize_decimal(pos.sell_amount),
            serialize_decimal(pos.sell_commission),
            serialize_decimal(pos.sell_exchange_rate),
            serialize_decimal(pos.profit),
            serialize_decimal(pos.income_pln),
            serialize_decimal(pos.costs_pln),
        ]

    for position in sorted(profit_positions, key=lambda x: (x.isin, x.sell_date)):
        profit_ws.append(create_row(position))

    for position in sorted(loss_positions, key=lambda x: (x.isin, x.sell_date)):
        loss_ws.append(create_row(position))

    wb.save(file)

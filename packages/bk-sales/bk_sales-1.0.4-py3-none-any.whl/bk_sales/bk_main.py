#beki

import json
import os
from .bk_prices import prices
from openpyxl import Workbook, load_workbook

xl_file = '/storage/emulated/0/bk_return_sales/bk_return_sales.xlsx'
json_file = "/storage/emulated/0/data.json"



if not os.path.exists(xl_file):
    print("Excel file not found")
    exit()

wb = load_workbook(xl_file)
ws = wb["BEKI"]
ws2 = wb["FUN"]


value_to_cell_mapping = {
    'knorraio': 'L3',
    'knorrchicken': 'L4',
    'lbt70g': 'L5',
    'lbl70g': 'L6',
    'lbt150g': 'L7',
    'lbl150g': 'L8',
    'lbt20g': 'L9',
    'lbl20g': 'L10',
    'luxsc70g': 'L11',
    'luxst70g': 'L12',
    'luxsc150g': 'L13',
    'luxst150g': 'L14',
    'signal60g': 'L15',
    'signal140g': 'L16',
    'shacoc350ml': 'L17',
    'concoc350ml': 'L18',
    'shaavo350ml': 'L19',
    'conavo350ml': 'L20',
    'shacoc700ml': 'L21',
    'concoc700ml': 'L22',
    'shaavo700ml': 'L23',
    'conavo700ml': 'L24',
    'shacoc15ml': 'L25',
    'concoc15ml': 'L26',
    'sunlightbar_200g': 'L27',
    'sunlight_40g': 'L28',
    'sunlight90g': 'L29',
    'sunlight160g': 'L30',
    'sunlight500g': 'L31',
    'sunlight1kg': 'L32',
    'sunlight5kg': 'L33',
    'omo40g': 'L34',
    'omo100g': 'L35',
    'omo500g': 'L36',
    'omo1kg': 'L37',
    'omo3kg': 'L38'
}

def add_sales():
    if not os.path.exists(json_file):
        print("JSON file not found")
        exit()

    with open(json_file, "r") as f:
        try:
            data = json.load(f)
        except json.JSONDecodeError:
            print("Error decoding JSON file")
            exit()

    changes_made = False  
    for key, value in data.items():
        if key in value_to_cell_mapping:
            ws[value_to_cell_mapping[key]] = value
            changes_made = True

    # Save the workbook only if changes were made
    if changes_made:
        wb.save(xl_file)
        print("Sales data added to Excel.")
    else:
        print("No relevant sales data to update.")


def update_prices():
  
    row = 3
    col = 15 

    for key, value in prices.items():
        ws2.cell(row=row, column=col, value=key)  # Product name
        ws2.cell(row=row, column=col + 1, value=value)  # Product price
        row += 1 
    wb.save(xl_file)

    print(f"Prices updated in {xl_file}")




#if __name__ == "__main__":
#    add_sales()

import pandas as pd
import openpyxl as px


class ExcelSheetsConcat(object):
    """
    Cette classe est utilisée pour lire et concaténer des données à partir de feuilles Excel.
    """

    def __init__(self, file_path):
        self.wb = None
        self.file_path = file_path

    def read_file(self):
        """Charge le classeur Excel en utilisant openpyxl."""
        try:
            self.wb = px.load_workbook(self.file_path)
        except FileNotFoundError as e:
            raise FileNotFoundError(f"Fichier introuvable : {self.file_path}") from e
        return self.wb

    def sheet_names(self):
        """Retourne les noms des feuilles dans le classeur."""
        if self.wb is None:
            try:
                self.read_file()
            except FileNotFoundError as e:
                print(e)
                return []
        return self.wb.sheetnames

    def infos(self):
        """Retourne les informations des feuilles dans le classeur.
        ------------------------------------------------------
        """
        infos = {}
        for sheet in self.sheet_names():
            infos[sheet] = {
                "max_row": self.wb[sheet].max_row,
                "max_column": self.wb[sheet].max_column,
            }
        return infos

    def get_column_names(self, sheet, row=8):
        """Extrait les noms des colonnes à partir de la ligne spécifiée de la feuille.
        ------------------------------------------------------------
        Paramètres :
            sheet: L'objet feuille à partir duquel extraire les noms des colonnes.
            row: Le numéro de la ligne contenant les noms des colonnes.

        """
        column_labels = [
            sheet.cell(row=row, column=i).value for i in range(1, sheet.max_column + 1)
        ]
        return column_labels  # Retourne une liste plate

    def sheets_data(
        self,
        header: int = 8,
        drop_last=True,
        last_rows=2,
        drop_first=True,
        first_rows=2,
    ):
        """Lit et concatène les données de toutes les feuilles, en utilisant la ligne spécifiée pour les en-têtes.
        ------------------------------------------------------------------------------
        Paramètres :
            row: Le numéro de la ligne à utiliser comme en-tête (par défaut 8).
            header: Le numéro de la ligne à utiliser comme en-tête (par défaut 8).
            drop_last: Indique s'il faut supprimer les dernières lignes de chaque feuille (par défaut True).
            last_rows: Le nombre de lignes à supprimer à la fin de chaque feuille (par défaut 2).
            first_rows: Le nombre de lignes à supprimer au début de chaque feuille (par défaut 2).
            drop_first: Indique s'il faut supprimer les premières lignes de chaque feuille (par défaut True).
        """

        try:
            self.read_file()
        except FileNotFoundError as e:
            print(e)
            return pd.DataFrame()

        all_data = []
        sheet_names = self.sheet_names()
        if not sheet_names:
            return pd.DataFrame()

        # En supposant que toutes les feuilles ont la même structure, on récupère les colonnes de la première feuille
        column_labels = self.get_column_names(self.wb[sheet_names[0]], row=header)

        for sheet in sheet_names:
            # Lit les données, en utilisant row-1 car pandas est indexé à partir de 0 et openpyxl à partir de 1
            df = pd.read_excel(
                self.file_path,
                sheet_name=sheet,
                header=header - 1,  # Ajuste pour l'indexation à partir de 0 dans pandas
                engine="openpyxl",
            )
            # Assure que les colonnes sont correctement nommées (en cas de divergences)
            if drop_last:
                df = df.iloc[
                    :-last_rows
                ]  # Supprime les dernières lignes (généralement les totaux)
            if drop_first:
                df = df.iloc[
                    first_rows:
                ]  # Supprime les premières lignes (généralement les en-têtes)
            df.columns = column_labels
            all_data.append(df)

        data = pd.concat(all_data, ignore_index=True)
        return data

    def save(self, file_name):
        """Sauvegarde les données dans un nouveau fichier Excel."""
        self.sheets_data().to_excel(file_name, index=False)
        print(f"Données sauvegardées sous le nom {file_name}")

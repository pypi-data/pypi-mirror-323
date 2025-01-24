from openpyxl import load_workbook
from typing import Union, List

class Excel(object):
    '''
    Excel复杂加工包装
    文件名需要使用绝对路径，例如：c:\\tool\\file_name.xlsx
    sheet可以作为str或者int类型，str需要传入sheet的名字，int需要传入sheet的index

    '''
    def __init__(self, file: str , sheet_index: str|int, ) -> object:
        # load workbook
        self.workbook = load_workbook(file)
        # select workbook
        self.sheet = self.workbook[sheet_index]

    


    def read_data_cols(self, head_index: int, cols: Union[str, List[str]]) -> List[dict]:
        """
        从表格中获取特定列的数据。

        Args:
            head_index (int): 表头占用的行数，从该行开始读取数据。
            cols (Union[str, List[str]]): 需要获取的列名称。
                如果只需要获取一列，可直接传入字符串；若需获取多列，请传入字符串列表。

        Returns:
            List[dict]: 包含列数据的字典列表，每个字典的键为列名，值为对应行的数据。

        Raises:
            TypeError: 如果 `cols` 参数不是字符串或字符串列表，则抛出此异常。
            ValueError: 如果指定的列在表格中找不到，则抛出此异常。
        """
        if isinstance(cols, str):
            col_names = [cols]
        elif isinstance(cols, list):
            col_names = cols
        else:
            raise TypeError("不支持的类型，请提供字符串或列表。")

        # 查找每列的索引
        column_index = []
        for col in col_names:
            col_info = self.find_value(col)
            if not col_info:
                raise ValueError(f"列 '{col}' 在表格中未找到。")
            column_index.append({'col': col, 'col_index': col_info['col_index']})

        # 获取指定列的数据
        result = []
        max_col_idx = max(item['col_index'] for item in column_index)

        for row in self.sheet.iter_rows(min_row=head_index + 1, max_col=max_col_idx, values_only=True):
            obj = {}
            for col_index in column_index:
                col_idx = col_index['col_index'] - 1
                if col_idx < len(row):
                    obj[col_index['col']] = row[col_idx]
                else:
                    obj[col_index['col']] = None  # 防止索引越界
            result.append(obj)

        return result

        

    def find_value(self, search_value: str, ) -> object:
        row_index = 1
        for row in self.sheet.iter_rows(values_only=True):
            col_index = 1
            for cell in row:
                if cell == search_value:
                    return {'row_index':row_index, 'col_index':col_index}
                col_index = col_index + 1
            row_index = row_index + 1
        return {'row_index':None, 'col_index':None}

    def close(self):
        self.workbook.close()